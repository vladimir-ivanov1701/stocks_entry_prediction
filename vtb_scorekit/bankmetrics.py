"""
Модуль содержит функции и классы для проведения статистическх тестов.

Реализованы функции рассчитывающие Джини(Сомерс Ди), модифицировнный джини, Индекс стабильности популяции,
Информационный индекс IV, расчет доверительных интервалов бутстрепом для Джини и Модифицированного Джини, 
и другие тесты.

Классы модуля проводят тесты.
На вход тесту подяются:
 1. Датафрейм (для тество производящих сравнение - 2 датафрейма)
 2. Имена столбцов, участвующих в расче теста
 3. Псевдонимы имен (для вывода в отчетах)
 4. Границы тестов
На выходе у экземпляра класса - 3 стандантрых параметра:
 1.Название test.tname
 2.Список таблиц  test.tbls, при этом первый элемент списка test.tbls[0] - является результатом
   теста в формате выходного отчета (светофор)
 3.Список графиков test.grafls, при этом первый элемент списка test.grafls[0] - является главным графиком

Графики представленны в виде объектов figure matplotlib

Расчитанный тест может быть выведен в эксель с помощью функции
test_to_excel(test,wb)

"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn import linear_model
import scipy.stats as sts
import seaborn as sns
import openpyxl
from openpyxl.styles import Font,Border,Side,Alignment,PatternFill
from openpyxl.utils import get_column_letter as gkl
from scipy.stats import beta
from scipy.stats import norm
from io import BytesIO
import matplotlib.ticker as mtick



__author__ = 'Anton Babak, Elizaveta Kapustina'
__vesion__ = '2.8.0'

# Корпоративная расцветка

color_blue=[0,40/255,130/255]
color_cyan=[0,170/255,255/255]
color_red=[234/255,107/255,80/255]
color_yellow=[241/255,204/255,86/255]
color_salad=[214/255,224/255,141/255]
color_green=[120/255,180/255,151/255]
color_gray=[197/255,199/255,200/255]
color_lblue=[51/255,83/255,155/255]
color_medium_blue = [30/255, 75/255, 210/255]
color_light_blue = [204/255, 238/255, 255/255]


light_rating_dict = {'С': None, 'З': 1, 'Ж': 2, 'К': 3}
kpk_light_func = lambda signal: light_rating_dict[signal]


#   Основные функции модуля

def somersdi(undep, res, unidir = True):
    """
    Вход
    ---------------------------------
    undep: независимая переменная
    res: граунд тру
    unidir: параметр направления. По умолчанию True т.е. направления независимой переменной и целевой совпадают.
    False -- направления направление независимой переменной и целевой не совпадают

    :returns значение метрики, ошибка (распределение нормальное)
    """
    dir_coeff = 1 if unidir else -1

    a = np.array(pd.crosstab(undep,res))
    o = np.array([0,0,0])
    N = a.sum()
    for (i,j), n in np.ndenumerate(a):
        sn = a[:,j].sum()
        if n != 0:
            d = a[:i,:j].sum()+a[(i+1):,(j+1):].sum()-a[:i,(j+1):].sum()-a[(i+1):,:j].sum()
            o = np.vstack((o,[n,d,sn]))
    w = N**2 - (a.sum(axis=0)**2).sum()
    S = (o[1:,0]*o[1:,1]).sum()/w
    se = ((o[1:,0]*(o[1:,1]-S*(N-o[1:,2]))**2).sum())**0.5/w*2
    return dir_coeff * S, se



def somersconfint(undep,res,alfa=0.05,numb=1000):
    """
    Расчет доверительных интервалов методом бутстрепа
    --------------------------------------------
    undep: независимая переменная
    res: граунд тру
    alpha: пуровень значимости
    numb: количество итераций бутстрепа
    :returns нижнюю и верхнюю границы доверительного интервала
    """
    if type(undep) == list:
        n = len(undep)
    else:
        n = undep.shape[0]
    ls = []
    for i in range(numb):
        index1 = np.random.randint(0,n,size=n)
        we = somersdi(undep[index1],res[index1])[0]
        ls.append(we)
    ls.sort()
    return ls[round(alfa*numb/2)],ls[-round(alfa*numb/2)]



def mgini(undep, res, s_order='natural', try_numb=100):
    """
    Расчет модифицированного коэффициента джини
    ------------------------------------------
    undep: независимая переменная
    res: граунд тру
    s_order: (порядок сортировки) он имеет значение для тех элементов, которые имеют одинаковые graund true
    natural - без сортирировки обработка массива - как он есть. Возможные значения strait, back, mean, 'random'
    (прямой, обратный, замена на среднее, случайный).
    try_numb: в случае выбора случайного порядка, задается количество испытаний (по дефолту = 100)
    .
    :returns модифицированный коэффициент Джини
    """

    # check and get number of samples
    assert undep.shape == res.shape
    n_samples = res.shape[0]

    # sort rows on ground true column
    # (from largest to smallest)

    arr = np.array([res, undep]).transpose()
    true_order = arr[arr[:, 0].argsort()][::-1, 0]
    #     print(np.sum(true_order,axis=-1),true_order.shape)
    L_true = np.cumsum(true_order) / np.sum(true_order)
    L_ones = np.linspace(1 / n_samples, 1, n_samples)

    # count Lorents by prediction

    if s_order == 'natural':
        pred_order = arr[arr[:, 1].argsort()][::-1, 0]
        L_pred = np.cumsum(pred_order) / np.sum(pred_order)
    elif s_order == 'back':
        arr2 = arr[np.argsort(arr[:, 0])][::-1, :]
        pred_order = arr2[arr2[:, 1].argsort()][::-1, 0]
        L_pred = np.cumsum(pred_order) / np.sum(pred_order)
    elif s_order == 'strait':
        arr2 = arr[np.argsort(arr[:, 0])]
        pred_order = arr2[arr2[:, 1].argsort()][::-1, 0]
        L_pred = np.cumsum(pred_order) / np.sum(pred_order)
    elif s_order == 'mean':
        arr2 = arr[np.argsort(arr[:, 0])]
        arr3 = arr2[::-1, :]
        pred_order = arr2[arr2[:, 1].argsort()][::-1, 0]
        pred_order2 = arr3[arr3[:, 1].argsort()][::-1, 0]
        L_pred = np.cumsum(pred_order) / np.sum(pred_order)
        L_pred = (np.cumsum(pred_order2) / np.sum(pred_order2) + L_pred) / 2
    elif s_order == 'random':
        L_pred = np.zeros(arr.shape[0])
        for i in range(try_numb):
            np.random.shuffle(arr)
            pred_order = arr[arr[:, 1].argsort()][::-1, 0]
            L_pred = L_pred + np.cumsum(pred_order) / np.sum(pred_order) / try_numb

            # get Gini coefficients (area between curves)
    G_true = np.sum(L_ones - L_true)
    G_pred = np.sum(L_ones - L_pred)

    # normalize to true Gini coefficient
    return G_pred / G_true



def mginiconfint(undep,res,alfa=0.05,numb=300, s_order='natural'):
    """
    Расчет доверительных интервалов для модифицированного джини
    ----------------------------------------------
    undep: независимая переменная
    res: граунд тру
    alpha: уровень значимости
    numb: количество итераций бутстрепа
    s_order: (порядок сортировки) он имеет значение для тех элементов, которые имеют одинаковые graund true
    natural - без сортирировки обработка массива - как он есть. Возможные значения strait, back, mean, 'random'
    (прямой, обратный, замена на среднее, случайный).

    :returns нижнюю и верхнюю границы доверительного интервала
    """
    if type(undep) == list:
        n = len(undep)
    else:
        n = undep.shape[0]
    if isinstance(undep, pd.Series):
        undep = undep.values
    if isinstance(res, pd.Series):
        res = res.values

    ls = []
    for i in range(numb):
        index1 = np.random.randint(0,n,size=n)
        we = mgini(undep[index1], res[index1], s_order)
        ls.append(we)
    ls.sort()

    return ls[round(alfa*numb/2)],ls[-round(alfa*numb/2)]


def binom_interval(success, total, confint=0.95):
    """
    Вычисляет границы доверительного интервала, используя биномиальное распределение
    --------------------------------------
    success: количество 'удачных' испытаний
    total: общее количество испытаний
    confint: доверительный интервал

    :returns нижнюю и верхнюю границу доверительного интервала
    """
    quantile = (1 - confint) / 2.
    lower = beta.ppf(quantile, success, total - success + 1)
    upper = beta.ppf(1 - quantile, success + 1, total - success)
    return (lower, upper)


def binom_interval_corr(pd, confint=0.95, rho=0.15):
    """
    Вычисляет границы доверительного интервала с учетом корреляций между активами,
    используя аппроксимацию биномального распределения к нормальному
    --------------------------------------
    pd: предсказанный pd
    confint: доверительный интервал
    rho: коэффициент корреляции между активами
    :returns нижнюю и верхнюю границу доверительного интервала
    """
    alpha = (1 - confint)/2
    l_q = (sts.norm.ppf(pd) + rho**0.5 * sts.norm.ppf(alpha))/(1 - rho)**0.5
    u_q = (sts.norm.ppf(pd) + rho**0.5 * sts.norm.ppf(1 - alpha))/(1 - rho)**0.5
    return (sts.norm.cdf(l_q), sts.norm.cdf(u_q))


def psi(a, b, n=10, method="bins"):
    """
    Считает индекс стабильности популяции в случае дискретных и непрерывных факторов.
    В случае наличия пропусков в данных, они удаляются и выставляется код ошибки.
    В случае несовпадения уникальных значений в дискретном варианте расчет производится на совпадающих бинах
    и выставляется код ошибки.

    Код ошибки:
    0 – ошибок нету
    1 – уникальные значения выборок не совпадают
    2 – есть пропуски в выборках
    3 – уникальные значения выборок не совпадают и есть пропуски

    --------------------------------------
    a: значения из первой выборки
    b: значения из второй выборки
    n: число интервалов, если фактор непрерывный
    method: bins - расчет будет идти по бинам
            quantile - расчет будет идти по бакетам
    return: значение psi
            код ошибки
    """
    a = a.apply(lambda x: pd.to_numeric(x))
    b = b.apply(lambda x: pd.to_numeric(x))
    
    x = np.array(a)
    y = np.array(b)
    x.sort()
    y.sort()
    x_sum = x.shape[0]
    y_sum = y.shape[0]
    null_error = 0
    match_error = 0
    
    # Проверка на наллы
    if ((pd.isnull(x).mean() != 0) | (pd.isnull(y).mean() != 0)):
        x = x[~pd.isnull(x)]
        y = y[~pd.isnull(y)]
        null_error = 1
    
    # Проверка на пустые выборки
    assert x.shape[0] != 0, 'Одна из выборок пустая'
    assert y.shape[0] != 0, 'Одна из выборок пустая'
    
    # Проверка на метод
    assert method in ['bins', 'quantile'], 'Неправильно выбран метод'
    
    # Проверка на n
    assert type(n) is int, 'n должно быть int'
    
    if method == "bins":
        # считаем по дискретным значениям
        if set(x) != set(y):
            match_error = 1
        
        x = pd.DataFrame(pd.DataFrame(x.round(5))[0].value_counts())
        y = pd.DataFrame(pd.DataFrame(y.round(5))[0].value_counts())

        xy = x.merge(y, left_index=True, right_index=True)
        xy.columns = ['q', 'w']
        xy['psi'] = xy.apply(lambda f: (f.q / x_sum - f.w / y_sum) * np.log(f.q / x_sum * y_sum / f.w), axis=1)

        return [xy['psi'].sum(), null_error*np.power(2, 1) + match_error*np.power(2, 0)]
    
    if method == "quantile":
        # считаем по n перцентилям первого столбца
        y_before = 0
        x_before = 0
        psi = 0

        for i in range(1, n):
            # Может быть случай когда одно значение с большим весом чем размер бакета, тогда могут совпасть
            # границы и функция упадет.
            cut_off = x[int(x_sum / n * i)]  # считаем отсекающее значение
            x_curr = np.argmin(x < cut_off)
            y_curr = np.argmin(y < cut_off)
            k = (y_curr - y_before) / y_sum
            g = (x_curr - x_before) / x_sum
            if k != 0:
                psi += (g - k) * np.log(g / k)
            x_before = x_curr
            y_before = y_curr

        # добавляем последний бакет
        k = 1 - (y_before) / y_sum
        g = 1 - (x_before) / x_sum
        psi += (g - k) * np.log(g / k)
        
        return [psi, null_error*np.power(2, 1) + match_error*np.power(2, 0)]
    
    
def iv_bin(d):
    """
    Вычисляет метрику Information Value
    -----------------------------------
    Вход
    d: датафрейм из 2-х столбцов: первый - переменная, второй - флаг дефолта.
    Ограничение - второе распределение - бинарное [0,1]
    :returns значение метрики IV, ?
    """
    df1 = d.reset_index()
    ind, v, c = df1.columns

    out_iv = df1.groupby(v).agg({c: 'sum', ind: 'count'}).reset_index()
    n_empty = np.sum(out_iv[(out_iv[c] == 0) | (out_iv[c] == out_iv[ind])][ind])
    out_iv = out_iv[(out_iv[c] != 0) & (out_iv[c] != out_iv[ind])]

    out_iv['good'] = out_iv[ind] - out_iv[c]
    out_iv[ind] = out_iv[ind] / np.sum(out_iv[ind])
    out_iv[c] = out_iv[c] / np.sum(out_iv[c])
    out_iv['good'] = out_iv['good'] / np.sum(out_iv['good'])

    out_iv['iv'] = (out_iv['good'] - out_iv[c]) * np.log(out_iv['good'] / out_iv[c])

    return out_iv['iv'].sum(), n_empty


# Новые тесты для LGD моделей

def NRMSE(x, y, normalize=True, borders_list=None):
    """
    Вычисляет корень среднего значения квадрата ошибки. Используется для оценки
    точности смоделированных результатов LGD/EAD.
    Показатель нормируется на среднее фактическое значение для обеспечения сопоставимости
    величин и размерностей (если параметр normalize = True).
    -----------------------------------
    Вход
    x: фактические значения LGD/EAD
    y: смоделированные значения LGD/EAD
    normalize: флаг применения нормировки на среднее фактическое значение
    borders_list: границы желтого и красного сигналов, значения в массиве должны идти
    в возрастающем порядке (по умолчанию borders_list=[0.1, 0.2])

    :returns корень среднего значения квадрата ошибки, значение сигнала в формате ['К', 'Ж','З', 'C']
    """
    if borders_list == None:
        borders_list = [0.1, 0.2]

    assert len(borders_list) == 2, "len(borders_list) != 2"
    assert borders_list[0] < borders_list[1], "The values in borders_list must be in ascending order"
    assert x.shape[0] == y.shape[0], "0-dims don't match"

    mean = x.mean() if normalize else 1
    n = x.shape[0]

    nrmse = (((x - y) ** 2).sum() / (n - 1)) ** 0.5 / mean

    if normalize:
        signal = 'З' if nrmse <= borders_list[0] else 'К' if nrmse > borders_list[1] else 'Ж'
    else:
        signal = 'С'

    return nrmse, signal


def LossShortfall(lgd, olgd, ead, borders_list= None):
    """
    Вычисляет коэффициент LossShortfall
    ------------------------------------
    Вход
    lgd: np.array or pd.Series with shape (n,)
    olgd: np.array or pd.Series with shape (n,)
    eadd: np.array or pd.Series with shape (n,)

    :returns коэффициент LossShortfall, значение сигнала в формате ['К', 'Ж','З']
    """
    assert lgd.shape[0] == olgd.shape[0] == ead.shape[0], "0-dims don't match"
    borders_list = [0, -0.1, -0.2]
    LS_coeff = 1 - (lgd * ead).sum() / (olgd * ead).sum()
    if (LS_coeff <= 0) & (LS_coeff > -0.1):
        signal = 'З'
    elif (LS_coeff <= -0.1) & (LS_coeff > -0.2):
        signal = 'Ж'
    else:
        signal = 'К'
    return LS_coeff, signal



def plot_factor_dist(sample_dict, parameters, save_path='', width_coeff=-1, save=False, rd=-1):
    """
    Функция построения распределения значений факторов в различных выборках.
    -------------------------------------------------------------
    Вход
    sample_dict: словарь, содержащий выборки.
    parameters: список факторов.
    save_path: директория, в которую нужно сохранить графики.
    width_coeff: коэффициент, определяющий ширину бинов. Если width_coeff=-1 (значение по умолчанию), 
                то ширина задается автоматически.
    save: флаг сохранения графиков в указанную директорию.
    rd: параметр определяющий количество знаков после запятой для округления значений фактора. Если rd=-1,
    значения не округляются.
    """
    frame={}
    for fact in parameters:
        mx=-1
        f=0
        frame[fact]=pd.DataFrame()
        for key, tmp in sample_dict.items():
            frame[fact][key]=(tmp[fact].value_counts()/tmp[fact].shape[0]*100)
            if (f==0)|(frame[fact][key].max()>mx):
                mx=frame[fact][key].max()
                f=1
        frame[fact]=frame[fact].sort_index()
        if width_coeff != -1:
            width =width_coeff/frame[fact].shape[0]
        elif frame[fact].shape[0] == 2:
            width =1.2/frame[fact].shape[0]
        elif frame[fact].shape[0] == 3:
            width =2.4/frame[fact].shape[0]
        else:
            width =3.5/frame[fact].shape[0]
        ax=frame[fact].plot(kind='bar', width=width,figsize=(16,10), 
                            color=[color_red, color_yellow, color_salad, color_green])
        w=width 
        ax.set(xlim=(-w, frame[fact].shape[0]-(1-w)), ylim=(0, mx*1.1))
        x = frame[fact].index
        x = np.round(x,rd) if rd>=0 else x
        ax.set_xticklabels(list(x), fontsize=16, rotation = 'horizontal')
        plt.yticks(fontsize=16)
        fmt = '%.0f%%' # Format you want the ticks, e.g. '40%'
        ax.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))
        plt.legend(loc='upper right', framealpha=1, fontsize = 14)
        plt.grid(color='w')
        plt.title('{}'.format(fact), fontsize=16)
        if save:
            plt.savefig(f'{save_path}/{fact}.jpg')
        #plt.show()
        plt.close()
        
        
def plot_factor_logic(df, parameters, target, target_label, save_path='', save=False, rd=-1):
    """
    Функция построения графиков, отображающих логику разбиения факторов.
    -------------------------------------------------------------
    Вход
    df: датафрейм с выборкой.
    parameters: список факторов.
    target: целевая переменная.
    target_label: название распределения целевой переменной (DR).
    save_path: директория, в которую нужно сохранить графики.
    save: флаг сохранения графиков в указанную директорию.
    rd: параметр определяющий количество знаков после запятой для округления значений фактора. Если rd=-1,
    значения не округляются.
    """
    frame={}

    for fact in parameters:
        frame0=df[[fact, target]]
        frame[fact]=frame0.groupby(fact, axis=0).sum()
        dr=[]
        shape=[]
        label=[]
        binscore=frame0[fact].astype(str)+'('+frame0[fact].astype(str)+')'
        frame0[fact+'_binscore']=binscore
        bns=binscore.unique()
        bns.sort()
        for bn in bns:
            z, o = float(bn[:-1].split('(')[0]),float(bn[:-1].split('(')[1])
            frame[fact].at[(o,z), target_label]=(100*frame0[(frame0[fact+'_binscore']==bn)&(frame0[target]==1)].shape[0]/frame0[frame0[fact+'_binscore']==bn].shape[0])
            frame[fact].at[(o,z),'shape']=(frame0[frame0[fact+'_binscore']==bn].shape[0]/frame0.shape[0]*100)
            frame[fact].at[(o,z), target]=(frame0[(frame0[fact+'_binscore']==bn)&(frame0[target]==1)].shape[0])

        frame[fact]=frame[fact].reset_index().sort_values(by=(fact)).reset_index()
        frame[fact][fact]=frame[fact][fact].astype(str)
        frame[fact]=frame[fact][[fact, target, 'shape', target_label]]
        fig, ax = plt.subplots(figsize=(10,7))
        width = frame[fact].shape[0]*0.125 # width of a bar

        frame[fact]['shape'].plot(ax=ax,kind='bar', width = width, label='Наблюдения', color=color_blue)
        plt.yticks(fontsize=13)
        fmt = '%.0f%%' # Format you want the ticks, e.g. '40%'
        ax.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))
        ax2 = ax.twinx()
        frame[fact][target_label].plot(ax=ax2, c=color_cyan, linewidth=3, marker='o', markersize=8)
        ax.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))
        labels=frame[fact][fact].tolist()
        labels = [np.round(float(i), rd) if rd>=0 else float(i) for i in labels]
        ax.set(xlim=(-width, len(labels)-(1-width)), ylim=(0, max(frame[fact]['shape'])*1.2))
        ax2.set(ylim=(0, max(frame[fact][target_label])*1.2))
#         ax.set_xticklabels(labels, fontsize=12, rotation=20)
        ax.set_xticklabels(labels, fontsize=14, rotation = 'horizontal')
        plt.yticks(fontsize=13)
        fmt = '%.1f%%'
        ax2.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))
        for ind in frame[fact].index:
            tmp=frame[fact].loc[ind]            
            plt.annotate(round(tmp[target_label], 1), xy=(ind, (tmp[target_label]+0.015)*1.02),ha='right',va='bottom',
                         color='k',backgroundcolor=color_gray, fontsize=14)
        plt.grid()
        plt.legend()
        plt.title('{}'.format(fact), fontsize=16)
        if save:
            plt.savefig(f'{save_path}/{fact}.jpg')
        #plt.show()
        plt.close()

#Классы модуля

# Новые классы
class CorrSpearman_test:
    def __init__(self, df, borders_list=None, **params):
        
        lgd_col = params['lgd_col']
        olgd_col = params['olgd_col']

        self.grafls = []
        self.tbls = []
        self.tbls.append(CorrSpearman_test.make_row(df, lgd_col, olgd_col, borders_list=borders_list)) # lgd_col, olgd_col
        

    @staticmethod
    def interval(df_corr, r_var, n, alpha=0.95):
        """
        Вход
        df_corr: датафрейм с коэффициентами корреляций Пирсона
        r_var: название столбца(фактора) по которому нужно вернуть коэф. кор. + границы доверительных интервалов
        n: размер выборки
        alpha: доверительный интервал

        returns: датафрейм с коэф.корр и доверительными интервалами
        """
        df1 = df_corr[[r_var]].copy()
        z = 0.5 * np.log(((1 + df1) / (1 - df1))[r_var].astype(float))
        se = 1 / (n - 3) ** 0.5
        #         chnge sts or stats
        zl, zu = sts.norm.interval(alpha)
        ci_l, ci_u = z + se * zl, z + se * zu
        df1['Нижняя граница {}% доверительного интервала'.format(alpha * 100)] = (np.exp(2 * ci_l) - 1) / (
        np.exp(2 * ci_l) + 1)
        df1['Верхняя граница {}% доверительного интервала'.format(alpha * 100)] = (np.exp(2 * ci_u) - 1) / (
        np.exp(2 * ci_u) + 1)
        return df1

    @staticmethod
    def make_table(df, cols1, cols2, alpha=0.95):
        """
        Вход
        df: датафрейм
        cols1: список с названиями переменных
        cols2: список с названиями переменных
        alpha: доверительный интервал

        returns датафрейм с коэф.корр и доверительными интервалами
        """
        df_corr = pd.DataFrame(columns=['Переменные',
                                        'Количество наблюдений',
                                        'Корреляция Спирмена',
                                        'Нижняя граница {}% доверительного интервала'.format(alpha * 100),
                                        'Верхняя граница {}% доверительного интервала'.format(alpha * 100)])
        rows = []
        for col1 in cols1:
            row = []
            for col2 in cols2:
                df_row, signal = CorrSpearman_test.make_row(df, col1, col2, alpha)
                df_corr = pd.concat([df_corr, df_row])
                row.append(signal)
            rows.append(row)
        return df_corr, pd.DataFrame(rows, columns=cols2, index=cols1)

    @staticmethod
    def make_row(df, col1, col2, alpha=0.95, borders_list=None):
        """
        Вход
        df: датафрейм
        cols1: название переменной
        cols2: название переменной
        alpha: доверительный интервал

        returns датафрейм (1, 5) с коэф.корр и доверительными интервалами
        """
        if borders_list is None:
            borders_list = [0.15, 0.3]
        
        df_corr = pd.DataFrame(index=[0],
                               columns=['Переменные',
                                        'Количество наблюдений',
                                        'Корреляция Спирмена',
                                        'Нижняя граница {}% доверительного интервала'.format(alpha * 100),
                                        'Верхняя граница {}% доверительного интервала'.format(alpha * 100),
                                        'Сигнал'])
        df_corr.iloc[0, 0] = '{}/{}'.format(col1, col2)
        df_corr.iloc[0, 2] = df[[col1, col2]].corr(method='spearman').iloc[0, 1]
        df_corr.iloc[0, 1] = df.shape[0]
        df_corr.iloc[0, 2:5] = CorrSpearman_test.interval(df_corr, df_corr.columns[2], df_corr.iloc[0, 1], alpha).iloc[0,:]
        x = df_corr.loc[0, 'Корреляция Спирмена']
        signal = 'К' if (x <= borders_list[0]) else 'Ж' if (x <= borders_list[1]) else 'З'
        df_corr.loc[0, 'Сигнал'] = signal
        return df_corr, signal

    
class LossShartfall_test:
    cols = ['Количество записей', 'Фактич.LGD', 'Прогнозное LGD', 'EAD',
            'Фактические потери', 'Прогнозные потери', 'Loss Shortfall', 'Сигнал']
    
    def __init__(self, df, weighted=True, border_list=None, **params):
#                  alias=None, short_alias=None, trust_level=0.95, unidir=True,
#                  sample_name='', df_ms=None, ms_col='', ms_col1='', ms_pd_col='' #col, cel
        
        df_dict = {'val': df}
        self.grafls = []
        self.tbls = []
        self.tbls.append(self.make_table(df_dict, weighted, border_list, **params)) # lgd_col, olgd_col, ead_col, ))
        

#     @staticmethod
    def make_table(self, df_dict, weighted=True, borders_list=None, **params): #lgd_col, olgd_col, ead_col, 
        """
        Вычисляет коэффициент LossShortfall

        Parameters
        ------------------------------------
        df_dict: словарь, где ключи -- названия выборок, а элементы датафреймы
        lgd_col: название столбца со значениями модельных lgd
        olgd_col: название столбца со значениями фактических lgd
        ead_col: название столбца со значениями ead
        weighted: если равен True, то модельный и фактические значения lgd взвешиваются на значения ead
        borders_list: список с границами светофоров

        Returns
        -------------------------------------
        df: датафрейм с результатами теста для каждой выборки из словаря
        """
        
        lgd_col = params['lgd_col']
        olgd_col = params['olgd_col']
        ead_col = params['ead_col']
        
        cols = LossShartfall_test.cols
        df = pd.DataFrame(columns=cols, index=df_dict.keys())

        for i, key in enumerate(df_dict):
            df.iloc[i, :] = LossShartfall_test.make_row(df_dict[key], lgd_col, olgd_col, ead_col, weighted=weighted,
                                                        borders_list=borders_list).iloc[0,:]
        return df

    @staticmethod
    def make_row(df1, lgd_col, olgd_col, ead_col, weighted=True, borders_list=None):
        """
        Вычисляет коэффициент LossShortfall

        Parameters
        ------------------------------------
        df1: датафрейм
        lgd_col: название столбца со значениями модельных lgd
        olgd_col: название столбца со значениями фактических lgd
        ead_col: название столбца со значениями ead
        weighted: если равен True, то модельный и фактические значения lgd взвешиваются на значения ead
        borders_list: список с границами светофоров

        Returns
        -------------------------------------
        df: датафрейм с результатами теста для выборки
        """
        if borders_list == None:
            borders_list = [0, 0.1]
#         assert borders_list[0] > borders_list[1], "borders_list[0] should be greater than borders_list[1]"
#         assert borders_list[2] == 0, "borders_list[2] should be equal to 0"
        cols = LossShartfall_test.cols
        df = pd.DataFrame(columns=cols)
        lgd = df1[lgd_col]
        olgd = df1[olgd_col]
        ead = df1[ead_col]
        ead = ead.apply(lambda x: pd.to_numeric(x))
        df.loc[0, [cols[3]]] = ead.sum()
        if not weighted:
            ead = np.ones(ead.shape)
        df[cols[0]] = lgd.shape[0]
        df[cols[1]] = (olgd * ead).sum() / ead.sum()
        df[cols[2]] = (lgd * ead).sum() / ead.sum()
        df[cols[4]] = df[cols[1]] * df[cols[3]]
        df[cols[5]] = df[cols[2]] * df[cols[3]]
        df[cols[6]] = 1 - df[cols[5]] / df[cols[4]]
        df[cols[7]] = df[cols[6]].apply(lambda x: 'К' if (x > borders_list[1]) else 
                                        'Ж' if ((x <= borders_list[1]) & (x > borders_list[0])) else 'З')
        
#         df[cols[7]] = df[cols[6]].apply(
#             lambda x: 'Ж' if ((x <= borders_list[0]) & (x > borders_list[1])) else 'З' if (
#                 (x <= borders_list[2]) & (x > borders_list[0])) else 'К')
        return df

class NRMSE_test:
    """
    Вычисляет корень среднего значения квадрата ошибки. Используется для оценки
    точности смоделированных результатов LGD/EAD.
    Показатель нормируется на среднее фактическое значение для обеспечения сопоставимости
    величин.
    """
    tname = 'Loss Shortfall'
    cols = ['Количество записей', 'Среднее фактическое значение', 'NRMSE', 'RMSE', 'Сигнал']
    borders_list = [0.1, 0.2]

    def __init__(self, df_dict=None, x_col=None, y_col=None, x=None, y=None, borders_list=None):
        """
        Вычисляет корень среднего значения квадрата ошибки. Используется для оценки
        точности смоделированных результатов LGD/EAD.
        Показатель нормируется на среднее фактическое значение для обеспечения сопоставимости
        величин.

        Parameters
        -----------------------------------
        1 вариант:
        Вход
        x: фактические значения LGD/EAD (pd.Series, pd.DataFrame with shape (n, 1), np.array)
        y: смоделированные значения LGD/EAD (pd.Series, pd.DataFrame with shape (n, 1), np.array)
        2 вариант:
        Вход
        df_dict: словарь, где ключи названия выборки, а элементы типа pd.DataFrame -- таблица с выборкой (dict)
        x_col: название столбца с фактическими значениями
        y_col: название столбца с модельными значениями

        borders_list: границы желтого и красного сигналов, значения в массиве должны идти в возрастающем порядке

        Note
        -----------------------------------
        при вызове параметра класса tbls[0] возвращает таблицу с результатами теста
        """
        self.tname = 'NRMSE'
        self.cols = ['Количество записей', 'Среднее фактическое значение', 'NRMSE', 'RMSE', 'Сигнал']
        self.tbls = []

        if borders_list == None:
            self.borders_list = [0.1, 0.2]

        if df_dict != None:
            self.df_dict = df_dict
            self.x_col = x_col
            self.y_col = y_col
            self.tbls.append(self.make_table_wrapper())
        elif x != None | y != None:
            self.x = x
            self.y = y
            self.tbls.append(self.make_row_wrapper())

    @staticmethod
    def make_table(df_dict, x_col, y_col, borders_list=None):
        """
        Вычисляет корень среднего значения квадрата ошибки. Используется для оценки
        точности смоделированных результатов LGD/EAD.
        Показатель нормируется на среднее фактическое значение для обеспечения сопоставимости
        величин.

        Parameters
        -----------------------------------
        df_dict: словарь, где ключи названия выборки, а элементы типа pd.DataFrame -- таблица с выборкой (dict)
        x_col: имя столбца с фактическими значениями LGD/EAD
        y_col: имя столбца с модельными значениями LGD/EAD
        borders_list: границы желтого и красного сигналов, значения в массиве должны идти в возрастающем порядке

        Returns
        -----------------------------------
        df: pd.DataFrame с результатами теста для каждой выборки
        """
        df = pd.DataFrame(columns=NRMSE_test.cols, index=df_dict.keys())
        for i, key in enumerate(df_dict):
            df.iloc[i, :] = NRMSE_test.make_row(df_dict[key][x_col], df_dict[key][y_col], borders_list).iloc[0, :]
        return df

    @staticmethod
    def make_row(x, y, borders_list=None):
        """
        Вычисляет корень среднего значения квадрата ошибки. Используется для оценки
        точности смоделированных результатов LGD/EAD.
        Показатель нормируется на среднее фактическое значение для обеспечения сопоставимости
        величин.

        Parameters
        -----------------------------------
        x: фактические значения LGD/EAD (pd.Series, pd.DataFrame with shape (n, 1), np.array)
        y: смоделированные значения LGD/EAD (pd.Series, pd.DataFrame with shape (n, 1), np.array)
        borders_list: границы желтого и красного сигналов, значения в массиве должны идти в возрастающем порядке

        Returns
        -----------------------------------
        df: строка типа pd.DataFrame с результатами теста
        """
        if borders_list == None:
            borders_list = NRMSE_test.borders_list
        assert len(borders_list) == 2, "len(borders_list) != 2"
        assert borders_list[0] < borders_list[1], "The values in borders_list must be in ascending order"
        assert x.shape[0] == y.shape[0], "0-dims don't match"

        df = pd.DataFrame(columns=NRMSE_test.cols)

        mean = x.mean()
        n = x.shape[0]
        nrmse = (((x - y) ** 2).sum() / (n - 1)) ** 0.5 / mean
        signal = 'З' if nrmse <= borders_list[0] else 'К' if nrmse > borders_list[1] else 'Ж'
        df.loc[0, :] = n, mean, nrmse, nrmse * mean, signal
        return df

    def make_row_wrapper(self, ):
        return self.make_row(self.x, self.y, self.borders_list)

    def make_table_wrapper(self, ):
        return self.make_table(self.df_dict, self.x_col, self.y_col, self.borders_list)

# -----------------------------------------------------------





# Базовый клас для метрик оценки предиктивной способности
# На вход: датафрейм, список переменных, целевая, список псеводнимов, список имен для графиков, границы, уровень доверия
# на выходе таблица - светофор и групповая фоторграфия.
class Iv_test:
    """
    Вычисляет коэффициент Information Value
    ------------------------------------
    Вход
    df: датафрейм
    col: список переменных
    cel: целевая переменная
    alias: список псевдонимов переменных
    short_alias: список имен переменных для графиков
    border_list: список с границами желтого и красного сигналов для каждой переменной.
    Пример задания списка - [[0.3, 0.1] for i in col]
    trust_level: доверительный интервал. По умолчанию 95%
    sample_name: Название выборки для подписи графиков.

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: список с графиками
    """
    def __init__(self, df, alias=None, short_alias=None, border_list=None, trust_level=0.95, sample_name='', **params): #col, cel
        col = params['col']
        self.cel = params['cel']
        
        self.df = df
#         self.cel = cel
        if type(col) == str:
            self.col = [col]
        else:
            try:
                self.col = [i for i in col]
            except:
                self.col = [col]

        if (alias==None) & (short_alias!=None):
            self.short_alias = short_alias
            self.alias = short_alias
        elif (alias!=None) & (short_alias==None):
            self.alias = alias
            self.short_alias = alias

        if (alias == None) & (short_alias == None):
            self.alias = self.col
            self.short_alias=self.col

        self.sample_name = sample_name
        self.trust_level = trust_level
        self.border_list = border_list
        if (border_list == None):
            self.border_list = [[0.3, 0.1] for i in col]

        self.tname = 'IV'
        self.value_list = []
        for var in col:
            ivi,ivi_empty = iv_bin(self.df[[var, self.cel]])
            self.value_list.append([ivi, ivi_empty/self.df.shape[0], ivi])
        self.tbls=[]
        self.grafls=[]
        self.tbls.append(self.make_table())
        self.grafls.append(self.make_big_plot())

    def make_table(self):
        df = self.df
        df_out = pd.DataFrame(columns=['Фактор/ модуль',
                                       self.tname,
                                       'Нижняя граница (' + str(int(self.trust_level * 100)) + '% дов. интервал)',
                                       'Верхняя граница (' + str(int(self.trust_level * 100)) + '% дов. интервал)',
                                       'Граница желтого сигнала',
                                       'Граница красного сигнала',
                                       'Сигнал'])

        for i in range(len(self.col)):
            som, low, hi = self.value_list[i]
            df_out.loc[df_out.shape[0], :-1] = [self.alias[i], float(som), float(low), float(hi),
                                                self.border_list[i][0],
                                                self.border_list[i][1]]
        df_out['Сигнал'] = df_out.apply(lambda x: 'З' if x[self.tname] > x['Граница желтого сигнала'] else
        'К' if x[self.tname] < x['Граница красного сигнала'] else 'Ж', axis=1)

        for i in range(1, 6):
            cl = df_out.columns[i]
            df_out[cl] = df_out[cl].apply(pd.to_numeric)

        return df_out

    def make_big_plot(self):
        fout = self.tbls[0]

        fig, ax = plt.subplots(1, 1, figsize=(9, 6))
        width = 0.7

        ax.bar(fout.index, fout[self.tname] * 100, yerr=abs(fout[fout.columns[2]] - fout[fout.columns[1]]) * 50,
               align='center', width=width, color=[0, 170 / 255, 255 / 255])

        ax.set_ylabel('IV, %%')
        ax.set_title(self.sample_name + '\n' + self.tname + ' для факторов модели')

        b_max = int(round(fout[self.tname].max() * 10 + 0.5) * 10 + 5)
        b_min = int(round(fout[self.tname].min() * 10 - 0.5) * 10 - 5)

        ax.set_xticks(fout.index)
        ax.set_xticklabels(self.short_alias, rotation='vertical')
        nvars = fout.shape[0]
        red_sight = [e[0] * 100 for e in self.border_list]
        yellow_sight = [e[1] * 100 for e in self.border_list]
        ax.plot(range(-1, nvars + 1), [red_sight[0]] + red_sight + [red_sight[-1]],
                color=[241 / 255, 204 / 255, 86 / 255])
        ax.plot(range(-1, nvars + 1), [yellow_sight[0]] + yellow_sight + [yellow_sight[-1]],
                color=[234 / 255, 107 / 255, 80 / 255])
        ax.grid(axis='x')
        plt.close(fig)
        return fig


class Gini_model_test(Iv_test):
    def __init__(self, df1, df2=None, name1='валидация', name2='разработка', alias=None, short_alias=None, 
                 border_list=None, unidir=False, kpk=False, **params): #, col1, col2, cel1, cel2,
        col1 = params['col']
        col2 = params['col']
        cel1 = params['cel']
        cel2 = params['cel']
        self.border_list = border_list
        if border_list and (len(border_list)>0):
            self.border_list = [border_list[0]]

        self.first_col = Gini_test(df1, alias, short_alias, self.border_list, unidir=unidir, kpk=kpk, **params)
        if df2 is not None:
            self.second_col = Gini_test(df2, alias, short_alias, self.border_list, unidir=unidir, **params)
        
        self.metric_name = self.first_col.tname
        self.tname = 'Относительный ' + self.metric_name
        self.short_alias = short_alias or alias or col1
        self.alias = alias or short_alias or col1
        self.name1 = name1
        self.name2 = name2
        
        if border_list and (len(border_list)>1):
            self.border_list = border_list[1]
        else:
            self.border_list = [0.05, 0.1]# for i in col1]
        
        self.tbls = []
        self.grafls = []
        
        df_out = self.first_col.tbls[0]
        
        if df2 is not None:
            df_out = self.make_table()
            
        self.tbls.append(df_out)
        
        if kpk:
            self.kpk_tbl = self.first_col.kpk_tbl
    
    def make_table(self):
        df_out = self.first_col.tbls[0]
        df2 = self.second_col.tbls[0]
        
        second_metric_name = self.metric_name + ' на выборке ' + self.name2
        df_out[second_metric_name] = df2[self.metric_name]#[df2.iloc[0][self.second_col.tname]]
        df_out['Относительное изменение'] = (df_out[self.metric_name] - df_out[second_metric_name]) / df_out[second_metric_name]
        
        df_out['Относительное изменение сигнал'] = df_out.apply(
            lambda x: 'З' if np.abs(x['Относительное изменение']) < self.border_list[0] else
            'К' if np.abs(x['Относительное изменение']) > self.border_list[1] else 'Ж', axis=1)

        return df_out

    
class Gini_test(Iv_test):
    """
    Вычисляет коэффициент Gini.
    Если у целевой переменной количество возможных значений больше 2-х, то считается метрика Сомерс ди.
    ------------------------------------
    Вход
    df: датафрейм
    col: список переменных
    cel: целевая переменная
    alias: список псевдонимов переменных
    short_alias: список имен переменных для графиков
    border_list: список с границами желтого и красного сигналов для каждой переменной (по умолчанию 10%, 15%),.
    Пример задания списка - [[0.15, 0.1] for i in col]
    trust_level: доверительный интервал. По умолчанию 95%
    unidir: флаг, что монотонность скор балла и вероятности дефолта имеют
    одинаковые направления (по умолчанию True, т.е. направления совпадают)
    sample_name: Название выборки для подписи графиков
    df_ms: датафрейм с PD по бакетам
    ms_col: столбец с рейтингами МШ в таблице df
    ms_col1: столбец с рейтингами МШ в таблице df_ms
    ms_pd_col: столбец с PD по бакетам в таблице df_ms

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: список с граффиками (графики WOE и ROC-кривая)
    """
    def __init__(self, df, alias=None, short_alias=None, border_list=None, trust_level=0.95, unidir=False,
                 sample_name='', df_ms=None, ms_col='', ms_col1='', ms_pd_col='', model_type = None, kpk=False, **params): #col, cel
        col = params['col']
        cel = params['cel']

        self.model_type = model_type
            
        if df_ms is not None:
            alias_, short_alias_ = None, None
            df_tmp = df.merge(df_ms, left_on=ms_col, right_on=ms_col1)
            df_tmp[ms_pd_col] = -df_tmp[ms_pd_col]
            col_ = col + [ms_pd_col]
            if alias is not None:
                alias_ = alias + ['model']
            if short_alias is not None:
                short_alias_ = short_alias + ['model']
        else:
            df_tmp, alias_, short_alias_, col_ = df, alias, short_alias, col,

        Iv_test.__init__(self, df_tmp, alias=alias_, short_alias=short_alias_,
                         border_list=border_list, trust_level=trust_level, sample_name=sample_name, 
                         **{'col': col_, 'cel': cel}) #, col_, cel
        self.sample_name = sample_name
        self.unidir = unidir

        if self.df[cel].value_counts().shape[0] == 2:
            self.tname = 'Джини'
        else:
            self.tname = 'Сомерс Ди'

        if (border_list == None):
            if self.model_type == 'LGD':
                self.border_list = [[0.1, 0.05] for i in self.col]
            else:
                self.border_list = [[0.15, 0.1] for i in self.col]
            if (df_ms is not None) or ((alias is not None) and (alias_[-1] == 'model')):
                self.border_list[-1] = [0.55, 0.4]
            elif (alias is not None) and (alias_[-1] == 'module'):
                self.border_list[-1] = [0.25, 0.1]
        else:
            self.border_list = border_list

        self.value_list = []
        for var in self.col:
            som, err = somersdi(self.df[var], self.df[self.cel], self.unidir)
            tl, tu = sts.norm.interval(self.trust_level)
            self.value_list.append((som, som + err * tl, som + err * tu))
        self.tbls = []
        self.grafls = []
        res_tbl = Iv_test.make_table(self)
        self.tbls.append(res_tbl)
#         self.grafls.append(Iv_test.make_big_plot(self))
#         self.grafls += self.make_woe_list()
#         self.grafls += self.make_roc_plot()

        if kpk:
            self.kpk_tbl = pd.DataFrame()
            model_res_tbl = res_tbl[res_tbl['Фактор/ модуль'] == 'model']
            if len(model_res_tbl) > 0:
                self.make_kpk_table(model_res_tbl, self.border_list[-1])
                
        
    def make_kpk_table(self, res_tbl, border_list):
        func_light_perc = lambda value: '{:.0%}'.format(value)
        func_value_perc = lambda value: '{:.2%}'.format(value)
        
        self.kpk_tbl = res_tbl.rename(columns={self.tname: 'value',
#                                                'ind': 'sample_cnt',
                                               'Нижняя граница (95% дов. интервал)': 'conf_int_95_min',
                                               'Верхняя граница (95% дов. интервал)': 'conf_int_95_max'
                                              })
        
        self.kpk_tbl['value'] *= 100
        self.kpk_tbl['conf_int_95_min'] *= 100
        self.kpk_tbl['conf_int_95_max'] *= 100
        self.kpk_tbl['threshold_red'] = f'<{func_light_perc(border_list[1])}'
        self.kpk_tbl['threshold_yellow'] = f'({func_light_perc(border_list[1])}; {func_light_perc(border_list[0])})'
        self.kpk_tbl['threshold_green'] = f'>= {func_light_perc(border_list[0])}'
        self.kpk_tbl['rating'] = self.kpk_tbl['Сигнал'].apply(kpk_light_func)
            
        self.kpk_tbl.drop(columns=['Сигнал', 'Фактор/ модуль', 'Граница желтого сигнала', 'Граница красного сигнала'], inplace=True)
        
    
    def get_df(self):
        return self.df

    def make_roc_plot(self):
        if self.tname == 'Джини':
            ls_out=[]
            i=0

            ascending = not self.unidir

            for var in self.col:
                cap_df = self.df[[var, self.cel]]
                segm = self.alias[i]
                col = 'blue'

                scr = cap_df.columns[0]
                flg = cap_df.columns[1]
                cap_df[scr] = cap_df[scr] - cap_df[scr].min()

                cap_df = cap_df.sort_values(scr, ascending=ascending).reset_index()
                cap_df['cumflg'] = cap_df[flg].cumsum()

                fig, ax = plt.subplots(1, 1, figsize=(7, 7))

                scrsum = cap_df.shape[0]
                flgsum = cap_df['cumflg'].max()
                ax.plot(cap_df.index / scrsum * 100, cap_df['cumflg'] / flgsum * 100, color=col, label='')
                ax.plot(np.linspace(0, 100, 100), np.linspace(0, 100, 100), color='y')

                ax.set_ylabel('% событий')
                ax.set_xlabel('% наблюдений')

                # ax.set_title('CAP кривая. ' + segm)
                ax.set_title(self.sample_name + '\n' + 'CAP кривая. ' + segm)
                i+=1
                ls_out.append(fig)
                plt.close(fig)
            return ls_out
        return None

    def make_woe_list(self):
        fout = self.tbls[0]
        df = self.df
        cel = self.cel

        ls = []

        for j in range(fout.shape[0]):
            vv = self.col[j]

            if df[vv].unique().shape[0] < 10:
                ls.append(self.make_woe_quant(vv))
            else:
                ls.append(self.make_woe_contin(vv, 5))

        return ls
    
    def make_woe_quant_list(self):
        fout = self.tbls[0]
        df = self.df
        cel = self.cel

        ls = []

        for j in range(fout.shape[0]):
            vv = self.col[j]

            if df[vv].unique().shape[0] < 10:
                ls.append(self.make_woe_quant(vv))

        return ls

    def make_woe_quant(self, vv):
        f_name = self.short_alias[self.col.index(vv)]
        cel = self.cel
        ddd = self.df[[vv, cel]].groupby(vv).agg({vv: 'count', cel: 'sum'})
        ddd.index.name = 'i1'
        ddd.reset_index(inplace=True)

        ddd['pd'] = ddd[cel] / ddd[vv]

        fig, ax = plt.subplots()
        ax2 = ax.twinx()
        s_norm = np.sum(ddd[vv]) / 100

        ax.bar(ddd.index, ddd[vv] / s_norm, align='center', width=0.5)
        ax2.plot(ddd.index, ddd.pd * 100, color='orange')
        ax.set_title(self.sample_name + '\n' + f_name)
        for r in ddd.iterrows():
            ax2.annotate(round(r[1].pd * 100, 2), xy=(r[0], r[1]['pd'] * 100))
        ax2.grid(None)
        ax.set_ylabel('% наблюдений')
        ax2.set_ylabel(' PD, %')
        ax2.set_ylim([0, ddd.pd.max() * 130])

        ax.set_xticks(ddd.index)
        ax.set_xticklabels(ddd.i1)
        plt.close(fig)
        return (fig, ddd.pd)

    def make_woe_contin(self, vv, n=5):
        df = self.df
        cel = self.cel
        f_name = self.short_alias[self.col.index(vv)]

        ddd0 = df[[vv, cel]].copy()
        ddd0[vv] = np.abs(ddd0[vv])
        dmax = ddd0[vv].max()
        dmin = ddd0[vv].min()

        ddd0['norm'] = (ddd0[vv] - dmin) / (dmax - dmin) * 100
        ddd0['n_baket'] = (ddd0['norm'] // n) / 100

        ddd1 = ddd0.groupby('n_baket').agg({vv: 'count', 'norm': 'mean', cel: 'mean'}).reset_index()
        ddd1 = ddd1.sort_values('norm')

        fig, ax = plt.subplots()
        ax2 = ax.twinx()

        ax.set_ylabel(f_name + ' , %')
        ax.plot(ddd1.index, ddd1.norm, color='green', label=f_name)
        
        ax2.grid(None)
        ax2.plot(ddd1.index, ddd1[cel], color='red', label='PD')
        ax2.set_ylabel(' PD, % ')
        ax.bar(ddd1.index, ddd1[vv] / ddd0.shape[0] * 100, align='center')
        ax.legend(loc=2)
        ax2.legend(loc=1)

        ax.set_xticklabels([])
        ax.set_title(self.sample_name + '\n' + f_name + ' и PD на фоне распределения.\n Бакеты по ' + str(n) + '%')
        plt.close(fig)
        return fig


class Mgini_test(Iv_test):
    """
    Вычисляет модифицированный коэффициент Gini.
    ------------------------------------
    Вход
    df: датафрейм
    col: список переменных
    cel: целевая переменная
    alias: список псевдонимов переменных
    short_alias: список имен переменных для графиков
    border_list: список с границами желтого и красного сигналов для каждой переменной (по умолчанию 10%, 15%),
    Пример задания списка - [[0.15, 0.1] for i in col]
    trust_level: доверительный интервал. По умолчанию 95%

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: список с графиками (графики WOE и ROC-кривая)
    """
    def __init__(self, df, alias=None, short_alias=None, border_list=None, trust_level=0.95, **params): #col, cel
        col = [params['lgd_col']]
        cel = params['olgd_col']
        alias = alias[-1:]
        
        if (border_list == None):
            self.border_list = [[0.1, 0.05] for i in col]
#             self.border_list = [[0.15, 0.1] for i in col]
            if (alias is not None) and (alias[-1] == 'module'):
                self.border_list[-1] = [0.25, 0.1]
            
#         Iv_test.__init__(self, df, col, cel, alias=alias, short_alias=short_alias,
#                          border_list=border_list, trust_level=trust_level)
        
        Iv_test.__init__(self, df, alias=alias, short_alias=short_alias,
                 border_list=self.border_list, trust_level=trust_level,
                 **{'col': col, 'cel': cel}) #, col_, cel

        self.tname = 'Модифицированный Джини'


        self.value_list = []
        for var in col:
            l, u = mginiconfint(self.df[var], self.df[self.cel], alfa=1 - self.trust_level, numb=300)
            self.value_list.append((mgini(self.df[var], self.df[self.cel], s_order='random'),
                                    l, u))

        self.tbls = []
        self.grafls = []
        self.tbls.append(Iv_test.make_table(self))
#         self.grafls.append(Iv_test.make_big_plot(self))


class Bin_test:
    """
    Проводит биномиальный тест
    Результат работы теста - таблица, графики в обычной и лог шкалах.
    Если нужен необычный график, его можно постротить отдельно. (см. функции модуля)
    -------------------------------------------------------------
    Вход
    df: датафрейм
    m0: столбец с обозначениями бакетов
    cel: столбец с фактическим PD
    pd0: столбец с модельным PD
    market: заголовок графика
    ct: CT для трансформации Байеса по умолчанию 1(без трансформации)
    start_baket: в случае если несколько первых бакетов объединяются в один.
    coloring: параметр, принимающий значения ['before', 'after'], определяет раскраску бакетов на графике.
       'before' - раскарска бакета по сигналу до корректировки (значение по умолчанию).
       'after' - раскараска бакетов по сигналу после корректировки.
    corr: значение коэффициента корреляции между дефолтами, если задан, проводится биномиальный тест с учетом дефолтов
    pit: флаг калибровки PIT. Если True, то на графике отрисовываются только уровни дефолта до корректировки и бакеты
        окрашиваются в цвет сигнала до корректировки. Если False на графике отрсовываются наблюдаемые скорректированные
        уровни дефолта, правило окраски бакетов задается параметром coloring.
    distrib: распределение для аппроксимации биномального.
       'norm' - нормальное (значение по умолчанию).
       'beta' - бета-распределение.
    Параметры класса
    self.tbls[0]: таблица с результатами биномиального теста
    self.tbls[1]: таблица с результатами теста на отклонение DR от среднего PD по разрядам рейтинговой шкалы - без учета корреляции 
    self.grafls: список с графиками
    """

    def __init__(self, df, market='', ct=1, start_baket='MA1', coloring='before', corr=0, pit=False,
                     distrib='norm', kpk=False, **params): #m0, cel, pd0, 
        m0 = params['m0']
        cel = params['cel']
        pd0 = params['pd0']
        
        self.market = market
        self.df = df
        self.m0 = m0
        self.cel = cel
        self.pd0 = pd0
        self.tname = 'Бин тест'
        self.ct = ct
        self.start_baket = start_baket
        self.di_95 = sts.norm.interval(0.95)[1]
        self.di_99 = sts.norm.interval(0.99)[1]
        self.coloring = coloring
        self.corr = corr
        self.pit = pit
        self.distrib = distrib

        self.tbls = []
        self.grafls = []
        self.tbls.append(self.make_table())
        
        if kpk:
            self.make_kpk_table()
        
        self.tbls[0].drop(columns=['default_cnt'], inplace=True)
                          
        if self.coloring == 'before':
            self.tbls.append(self.make_signal_table(self.tbls[0], 'signal before corr', 'perc'))
        elif self.coloring == 'after':
            self.tbls.append(self.make_signal_table(self.tbls[0], 'signal after corr', 'perc'))
              
        self.grafls.append(self.make_log_plot())
        self.grafls.append(self.make_main_plot())
        
        self.tbls[0].columns = ['Ранг Мастер шкалы', 'Количество наблюдений', 'Доля наблюдений',
                                'PD модели', 'Наблюдаемый уровень дефолта',
                                'Наблюдаемый уровень дефолта + преобразование', 'Сигнал (до корр.)',
                                'Сигнал (после корр.)',
                                'Нижняя граница (95% дов. интервал)',
                                'Верхняя граница (95% дов. интервал)', 'Нижняя граница (99% дов. интервал)',
                                'Верхняя граница (99% дов. интервал)']
                          

    def make_kpk_table(self):
        self.kpk_tbl = self.tbls[0].copy()[[self.m0, self.pd0, 'ind', 'default_cnt', 
                                            'l_err', 'u_err', 'l_err99', 'u_err99', 'signal before corr']]
        self.kpk_tbl = self.kpk_tbl.rename(columns={self.m0: 'rang_master_scale',
                                               self.pd0: 'value',
                                               'ind': 'sample_cnt',
                                               'l_err': 'conf_int_95_min',
                                               'u_err': 'conf_int_95_max',
                                               'l_err99': 'conf_int_99_min',
                                               'u_err99': 'conf_int_99_max'
                                              })
        self.kpk_tbl['threshold_red'] = 'Непопадание в 99% ДИ'
        self.kpk_tbl['threshold_yellow'] = 'Попадание в 99%, непопадание в 95% ДИ'
        self.kpk_tbl['threshold_green'] = 'Попадание в 95%'

        self.kpk_tbl['rating'] = self.kpk_tbl['signal before corr']
        self.kpk_tbl.drop(columns=['signal before corr'], inplace=True)
        

    def make_table(self):
        df = self.df.copy()
        m0 = self.m0
        if self.start_baket != 'MA1':
            df[m0] = df[m0].apply(lambda x: self.start_baket if x < self.start_baket  else x)

        c0 = self.cel
        pd0 = self.pd0
        m0 = self.m0
        ci_95 = sts.norm.interval(0.95)[1]

        df_rang = (df.groupby(m0).agg({c0: 'sum', pd0: ['mean', 'count']}))
        df_rang.columns = df_rang.columns.droplevel(0)
        df_rang = df_rang.reset_index()
        df_rang.rename(columns={'sum': c0, 'mean': pd0, 'count': 'ind'}, inplace=True)
        df_rang['default_cnt'] = df_rang[c0]
        df_rang[c0] = df_rang[c0] / df_rang['ind']
        df_rang['c0'] = df_rang[c0]  # костыль для сигнала до корректировки
        df_rang['err'] = ((df_rang[pd0] * (1 - df_rang[pd0]) / df_rang['ind']) ** 0.5)

        if self.ct != 1:
            dr = (df_rang[c0] * df_rang['ind']).sum() / df_rang['ind'].sum()
            df_rang['c2'] = df_rang[c0].apply(
                lambda x: x * self.ct / dr / (x * self.ct / dr + (1 - x) * (1 - self.ct) / (1 - dr)))
        else:
            df_rang['c2'] = df_rang[c0]

        for indx in df_rang.index:
            tmp = df_rang.loc[indx]
            meanpd = tmp[pd0]
            cnt = tmp['ind']
            if self.corr:
                lower99, upper99 = np.array(self.binom_interval_corr(tmp[pd0], confint=0.99, rho=self.corr))
                lower95, upper95 = np.array(self.binom_interval_corr(tmp[pd0], confint=0.95, rho=self.corr))
            else:         
                if self.distrib == 'beta':
                    min_r, max_r = np.array(self.beta_binom_interval(np.round(meanpd * cnt), cnt, confint=0.99)) 
                    min_y, max_y = np.array(self.beta_binom_interval(np.round(meanpd * cnt), cnt, confint=0.95)) 

                elif self.distrib == 'norm':        
                    min_r = max(0, meanpd - np.sqrt(meanpd * (1 - meanpd) / cnt) * sts.norm.interval(0.99)[1]) 
                    min_y = max(0, meanpd - np.sqrt(meanpd * (1 - meanpd) / cnt) * sts.norm.interval(0.95)[1]) 

                    max_y = (meanpd + np.sqrt(meanpd * (1 - meanpd) / cnt) * sts.norm.interval(0.95)[1]) 
                    max_r = (meanpd + np.sqrt(meanpd * (1 - meanpd) / cnt) * sts.norm.interval(0.99)[1]) 
                      
            df_rang.at[indx, 'u_err'] = max_y
            df_rang.at[indx, 'u_err99'] = max_r
            df_rang.at[indx, 'l_err'] = max(min_y, 0)
            df_rang.at[indx, 'l_err99'] = max(min_r, 0)
            
            df_rang.at[indx, 'signal after corr'] = self.flag_check(tmp.c2, min_y, max_y, min_r, max_r)
            df_rang.at[indx,'signal before corr'] = self.flag_check(tmp.c0, min_y, max_y, min_r, max_r)

        df_rang['perc'] = df_rang['ind'] / (df_rang['ind'].sum())

        return df_rang.loc[:,[m0, 'ind', 'perc', 'default_cnt', pd0, c0, 'c2', 
                              'signal before corr', 'signal after corr', 
                              'l_err', 'u_err', 'l_err99', 'u_err99']]
    
    def make_signal_table(self, df, signal, perc):
        col1 = 'Доля набл. в сигнальных и граничных бакетах'
        df_res = df.groupby([signal]).agg({perc: ['sum', 'count']})
        df_res.columns = df_res.columns.droplevel(0)

        if 'Н Ж' not in df_res.index:
            df_res.loc['Н Ж'] = 0, 0
        if 'Н К' not in df_res.index:
            df_res.loc['Н К'] = 0, 0

        df_res.loc[col1, df_res.columns] = df_res.loc['Н Ж', 'sum'] + df_res.loc['Н К', 'sum'] , 0
        df_res.loc['Сигнал', df_res.columns] = 'З' if df_res.loc[col1, 'sum'] < 0.2 \
                                                    else 'К' if df_res.loc['Н К', 'sum'] >= 0.2 \
                                                    else 'Ж', ''
        return df_res

    def make_main_plot(self):
        df_rang = self.tbls[0]
        m0, ind, perc, pd0, c0, c2, signal_before, signal_after, l_err, u_err, l_err99, u_err99 = df_rang.columns

        fig, ax = plt.subplots(1, 1, figsize=(10, 7))

        ax2 = ax.twinx()
        ax.bar(df_rang.index - 0.0, df_rang[ind] / df_rang[ind].sum() * 100, color='grey')

        l = ['С', 'П Ж', 'П К', 'Н Ж', 'Н К', 'Серый']
        colors = [color_green, color_yellow, color_yellow, color_yellow, color_red, color_gray]

        if self.coloring == 'before' or self.pit:
            signal = signal_before
        elif self.coloring == 'after':
            signal = signal_after
        else:
            raise ValueError('The value of variable coloring is not in [\'before\', \'after\']')


        MS1 = df_rang[[perc, signal]]
        # width = MS1.shape[0] * 0.05
        for i, c in zip(l, colors):
            MS1[i] = MS1[perc][MS1[signal] == i]
            ax.bar(MS1.index - 0.0, (MS1[i] * 100), color=c)

        ax.grid(None)
        ax2.grid(None)

        ax2.plot(df_rang.index, (df_rang[l_err99]), color='r', label='')
        ax2.plot(df_rang.index, (df_rang[u_err99]), color='r', label='границы 99% Дов. Инт.')


        ax2.scatter(df_rang.index, (df_rang[c0]), color='blue', label='Наблюдаемые дефолты')
        if not self.pit:
            ax2.scatter(df_rang.index, (df_rang.c2), color='orange',
                        label='Байесовский сдвиг на величину ЦТ={:.3f}%'.format(self.ct * 100))

        ax.set_title('Наблюдаемый и предсказываемый уровень дефолтов на фоне распределения по бакетам МШ.\n '
                     + self.market)
        ax.set_xticks(df_rang.index)
        ax.set_xticklabels(df_rang[m0], rotation='vertical')

        ax.set_ylabel('% наблюдений')
        ax2.set_ylabel('PD')

        # ax2.set_yticklabels([0 if x < 0 else '' for x in ax2.get_yticks()])
        fmt = '%.01f%%'
#         fmt = '%.0f%%'
        ax2.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))
        ax2.hlines((0), -0.0, df_rang.shape[0], color='white', linestyles='dashed', alpha=0)

        ax2.set_ylim([0, df_rang.c2.max() * 1.2])
        ax2.legend(loc=2)
        plt.close(fig)
        return fig

    def make_log_plot(self):
        df_rang = self.tbls[0]
        m0, ind, perc, pd0, c0, c2, signal_before, signal_after, l_err, u_err, l_err99, u_err99 = df_rang.columns

        fig, ax = plt.subplots(1, 1, figsize=(10, 7))

        ax2 = ax.twinx()
        ax.bar(df_rang.index , df_rang[ind] / df_rang[ind].sum() * 100, color='grey')

        l = ['С', 'П Ж', 'П К', 'Н Ж', 'Н К', 'Серый']
        colors = [color_green, color_yellow, color_yellow, color_yellow, color_red, color_gray]

        if self.coloring == 'before' or self.pit:
            signal = signal_before
        elif self.coloring == 'after':
            signal = signal_after
        else:
            raise ValueError('The value of variable coloring is not in [\'before\', \'after\']')

        MS1 = df_rang[[perc, signal]]
        # width = MS1.shape[0] * 0.05
        for i, c in zip(l, colors):
            MS1[i] = MS1[perc][MS1[signal] == i]
            ax.bar(MS1.index , (MS1[i] * 100), color=c)

        ax.grid(None)
        ax2.grid(None)

        ax2.plot(df_rang.index, (df_rang[l_err99]) * 100, color='r', label='')
        ax2.plot(df_rang.index, (df_rang[u_err99]) * 100, color='r', label='границы 99% Дов. Инт.')

        ax2.scatter(df_rang.index, (df_rang[c0]) * 100, color='blue', label='Наблюдаемые дефолты')
        if not self.pit:
            ax2.scatter(df_rang.index, (df_rang.c2) * 100, color='orange',
                        label='Байесовский сдвиг на величину ЦТ={:.3f}%'.format(self.ct * 100))
        plt.yscale('log')

        ax.set_title(
            'Наблюдаемый и предсказываемый уровень дефолтов на фоне распределения по бакетам МШ.'
            '\nЛогарифмическая шкала PD.\n'
            + self.market)
        ax.set_xticks(df_rang.index)
        ax.set_xticklabels(df_rang[m0], rotation='vertical')

        ax.set_ylabel('% наблюдений')
        ax2.set_ylabel('PD')
        fmt = '%.01f%%'
#         fmt = '%.0f%%'
        ax2.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))
        ax2.hlines((1), -0, df_rang.shape[0], color='white', linestyles='dashed', alpha=0.8)
        ax2.hlines((10), -0.0, df_rang.shape[0], color='white', linestyles='dashed', alpha=0.8)

        # ax2.set_yticklabels([0 if x < 0 else '' for x in ax2.get_yticks()])

        ax2.set_ylim([0.01, df_rang.c2.max() * 100 * 1.2])
        ax2.legend(loc=2)
        plt.close(fig)
        return fig

    def flag_check(self, dr_corr, min_y, max_y, min_r, max_r):
        answer = ''
        if dr_corr == 0 or np.isnan(dr_corr):
            answer = 'Серый'
        elif (min_y < dr_corr < max_y):
            answer = 'С'
        elif min_r < dr_corr < min_y:
            answer = 'П Ж'
        elif dr_corr < min_r:
            answer = 'П Ж'
        elif max_y < dr_corr < max_r:
            answer = 'Н Ж'
        elif dr_corr>max_r:
            answer = 'Н К'
        return answer
    
    def beta_binom_interval(self, success, total, confint=0.95):
        """
        Вычисляет границы доверительного интервала, используя биномиальное распределение
        --------------------------------------
        success: количество 'удачных' испытаний
        total: общее количество испытаний
        confint: доверительный интервал

        :returns нижнюю и верхнюю границу доверительного интервала
        """
        quantile = (1 - confint) / 2.
        lower = beta.ppf(quantile, success, total - success + 1)
        upper = beta.ppf(1 - quantile, success + 1, total - success)
        return (lower, upper)


    def binom_interval_corr(self, pd, confint=0.95, rho=0.15):
        """
        Вычисляет границы доверительного интервала с учетом корреляций между активами,
        используя аппроксимацию биномального распределения к нормальному
        --------------------------------------
        pd: предсказанный pd
        confint: доверительный интервал
        rho: коэффициент корреляции между активами
        :returns нижнюю и верхнюю границу доверительного интервала
        """
        alpha = (1 - confint)/2
        l_q = (sts.norm.ppf(pd) + rho**0.5 * sts.norm.ppf(alpha))/(1 - rho)**0.5
        u_q = (sts.norm.ppf(pd) + rho**0.5 * sts.norm.ppf(1 - alpha))/(1 - rho)**0.5
        return (sts.norm.cdf(l_q), sts.norm.cdf(u_q))
      


class Bin_test_kpk:
    def __init__(self, df, ct=1, kpk=False, bayes_check=False, bayes_flg=False, **params):
        cel = params['cel']
        
        self.tbls = []
        self.grafls = []
        
        if type(ct) is list:
            bin_tbl = self.module_merge(df, ct, params)
            bintest = Bin_test(df, kpk=kpk, **params)
        else:
            dr = df[cel].mean()
            bayes_flg = False if np.round(dr, 4) == ct else True
            
            bintest = Bin_test(df, ct=ct, kpk=kpk, **params)
            bin_tbl = bintest.tbls[0]
        
        bin_tbl_with_k, grf = self.binomial_test_new_with_k(bin_tbl, bayes_check=bayes_check, bayes_flg=bayes_flg)
        DR_corr = (bin_tbl_with_k['DR + преобразование'] * bin_tbl_with_k['Количество наблюдений']).sum() \
            / bin_tbl_with_k['Количество наблюдений'].sum()
        print('DR correction:', DR_corr)
        self.tbls.append(bin_tbl_with_k)
        self.grafls.append(grf)
            
        if kpk:
            self.kpk_tbl = bintest.kpk_tbl
            self.make_kpk_table(self.tbls[0])#[[self.m0, self.pd0, 'ind', 'default_cnt']])
        

    def make_kpk_table(self, res_tbl):
        self.kpk_tbl.set_index('rang_master_scale', inplace=True)
        self.kpk_tbl.drop(columns = ['value', 'conf_int_95_min', 'conf_int_95_max', 'conf_int_99_min', 'conf_int_99_max', 'rating'], inplace=True)
        self.kpk_tbl[['value', 'conf_int_95_min', 'conf_int_95_max', 
                  'conf_int_99_min', 'conf_int_99_max']] = res_tbl[['k~', 'k5%', 'k95%', 'k1%', 'k99%']]

        self.kpk_tbl['rating'] = res_tbl['Сигнал']
            
    def module_merge(self, df, ct, params):
        tmp_mod=df[df['module']==0]
        binom=Bin_test(tmp_mod, market='Модель', ct=ct[0], **params)
        table=binom.tbls[0].round(8)

        table=table[['Ранг Мастер шкалы','Количество наблюдений','Наблюдаемый уровень дефолта','Наблюдаемый уровень дефолта + преобразование','PD модели']]

        mods=[1,2,3]
        drs = ct[1:]
        for mod, dr in zip(mods, drs):

            tmp_mod=df[df['module']==mod]
            binom=Bin_test(tmp_mod, market='Модель', ct=dr, **params)
            table_new=binom.tbls[0].round(8)
            table_new=table_new[['Ранг Мастер шкалы','Количество наблюдений','Наблюдаемый уровень дефолта','Наблюдаемый уровень дефолта + преобразование','PD модели']]

            table=table.merge(table_new, on=['Ранг Мастер шкалы', 'PD модели'], how='outer').fillna(0)
            table['dflts']=table['Наблюдаемый уровень дефолта + преобразование_x']*table['Количество наблюдений_x']+table['Наблюдаемый уровень дефолта + преобразование_y']*table['Количество наблюдений_y']
            table['dflts_bf']=table['Наблюдаемый уровень дефолта_x']*table['Количество наблюдений_x']+table['Наблюдаемый уровень дефолта_y']*table['Количество наблюдений_y']

            table['Количество наблюдений']=table['Количество наблюдений_x']+table['Количество наблюдений_y']
            table['Наблюдаемый уровень дефолта + преобразование']=table['dflts']/table['Количество наблюдений']
            table['Наблюдаемый уровень дефолта']=table['dflts_bf']/table['Количество наблюдений']
            table=table[['Ранг Мастер шкалы','Количество наблюдений','Наблюдаемый уровень дефолта',
                         'Наблюдаемый уровень дефолта + преобразование','PD модели']]

        table=table.sort_values(by=['Ранг Мастер шкалы'])
        table['Доля наблюдений']=table['Количество наблюдений']/table['Количество наблюдений'].sum()
        return table
        
    def binomial_test_new_with_k(self,
        MS,
        plt_title='', 
        xlsx_title=None,
        png_title=None,
        distrib='norm', 
        draw_with_corr=True, 
        use_log_scale=True,
        bayes_check=False,
        bayes_flg=True
    ):

        cols_to_return = ['Ранг Мастер шкалы', 'Количество наблюдений', 'Доля наблюдений',
                          'PD модели', 'DR', 'DR + преобразование', 
#                           'Наблюдаемый уровень дефолта',
#                           'Наблюдаемый уровень дефолта + преобразование', 
#                           'Сигнал (до корр.)', 'Сигнал (после корр.)',
                          'k5%', 'k95%', 'k1%', 'k99%', 'DR~', 'k~',
                          'Сигнал']
        
        alpha=0.05
        MS['m1']=norm.ppf(1-alpha)
        MS['m2']=MS['Количество наблюдений']*MS['PD модели']*(1-MS['PD модели'])
        MS['m3']=MS['Количество наблюдений']*MS['PD модели']
        MS['k5%']=MS['m1']*np.sqrt(MS['m2'])+MS['m3']

        alpha=0.01
        MS['m1']=norm.ppf(1-alpha)
        MS['m2']=MS['Количество наблюдений']*MS['PD модели']*(1-MS['PD модели'])
        MS['m3']=MS['Количество наблюдений']*MS['PD модели']
        MS['k1%']=MS['m1']*np.sqrt(MS['m2'])+MS['m3']

        alpha=0.95
        MS['m1']=norm.ppf(1-alpha)
        MS['m2']=MS['Количество наблюдений']*MS['PD модели']*(1-MS['PD модели'])
        MS['m3']=MS['Количество наблюдений']*MS['PD модели']
        MS['k95%']=MS['m1']*np.sqrt(MS['m2'])+MS['m3']

        alpha=0.99
        MS['m1']=norm.ppf(1-alpha)
        MS['m2']=MS['Количество наблюдений']*MS['PD модели']*(1-MS['PD модели'])
        MS['m3']=MS['Количество наблюдений']*MS['PD модели']
        MS['k99%']=MS['m1']*np.sqrt(MS['m2'])+MS['m3']

        MS['DR~']=np.maximum(MS['Наблюдаемый уровень дефолта + преобразование'],MS['Наблюдаемый уровень дефолта'])
        if bayes_check:
            print('Bayes check!')
            if bayes_flg:
                MS['DR~']=MS['Наблюдаемый уровень дефолта + преобразование']
            else:
                MS['DR~']=MS['Наблюдаемый уровень дефолта']

        MS['k~']=MS['DR~']*MS['Количество наблюдений']

        for gr in  MS['Ранг Мастер шкалы'].unique():
            y_max = MS[MS['Ранг Мастер шкалы'] == gr]['k5%'].iloc[0]
            r_max = MS[MS['Ранг Мастер шкалы'] == gr]['k1%'].iloc[0]
            dr_corr2 = MS[MS['Ранг Мастер шкалы'] == gr]['k~'].iloc[0]
            cou = MS[MS['Ранг Мастер шкалы'] == gr]['Количество наблюдений'].iloc[0]
            answer = 'Серый'
            if (y_max < dr_corr2 < r_max)&(dr_corr2 != 0):
                answer = 'Ж'
            elif cou == 0:
                answer = 'Серый'
            elif (dr_corr2 > r_max) &(dr_corr2 != 0):
                answer = 'К'
            elif (dr_corr2 <= y_max)&(dr_corr2 != 0):
                answer = 'С'
            MS.loc[MS['Ранг Мастер шкалы']==gr, 'Новый сигнал по алгоритму ЦБ'] = answer


        if xlsx_title:
            MS.to_excel(xlsx_title, index=None)

        MS1 = MS # MS это табличка из сценария, по которой отрисовывается график калибровки

        l = ['С', 'П Ж', 'П К', 'Н Ж', 'Н К', 'К', 'Серый', 'Ж']
        colors = [color_green, color_yellow, color_red, color_yellow, color_red, color_gray]

        for i in l:
            if draw_with_corr:
                MS1[i] = MS['Доля наблюдений'][MS['Новый сигнал по алгоритму ЦБ'] == i]
            else:
                MS1[i] = MS['Доля наблюдений'][MS['Новый сигнал по алгоритму ЦБ'] == i]
        MS1['rang']=MS1['Ранг Мастер шкалы']
        MS1=MS1.set_index('Ранг Мастер шкалы')
        MS1=MS1.rename(columns={'rang':'Ранг Мастер шкалы'})
        fig, ax = plt.subplots(figsize=(10,6)) #обычный
        width = MS1.shape[0]*0.05

        MS1['Н Ж'].plot(ax=ax, kind='bar', color=color_yellow)
        plt.yticks(fontsize=12)

        MS1['Н К'].plot(ax=ax, kind='bar', color=color_red)
        plt.yticks(fontsize=12)

        MS1['С'].plot(ax=ax, kind='bar', label='Зеленый сигнал по бакету', color=color_green)
        plt.yticks(fontsize=12)

        MS1['Серый'].plot(ax=ax, kind='bar', label='Серый сигнал по бакету', color=color_gray)
        plt.yticks(fontsize=12)

        MS1['П Ж'].plot(ax=ax, kind='bar', label='Желтый сигнал по бакету', color=color_yellow)
        plt.yticks(fontsize=12)

        MS1['Ж'].plot(ax=ax, kind='bar', label='Желтый сигнал по бакету', color=color_yellow)
        plt.yticks(fontsize=12)

        MS1['П К'].plot(ax=ax, kind='bar', label='Красный сигнал по бакету', color=color_red)
        plt.yticks(fontsize=12)

        MS1['К'].plot(ax=ax, kind='bar', label='Красный сигнал по бакету', color=color_red)
        plt.yticks(fontsize=12)

        fmt = '%.0f%%' # Format you want the ticks, e.g. '40%'
        #ax.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))

        plt.legend(['Желтый сигнал по бакету', 
                    'Красный сигнал по бакету',
                    'Зеленый сигнал по бакету',
                    'Серый сигнал по бакету'], frameon = False, fontsize = 'small', loc = (0.008, 0.75))

        ax2 = ax.twinx()
        if use_log_scale:
            ax2.set_yscale('log')

        MS1['k~'].plot(ax=ax2, c=color_medium_blue, linewidth=4, marker='o', markersize=4,label = '$\widetilde{k}$')
        MS1['k5%'].plot(ax=ax2, c=color_yellow, linewidth=1, marker='o', markersize=4, label = 'k5%' )
        MS1['k1%'].plot(ax=ax2, c=color_red, linewidth=1, marker='o', markersize=4, label = 'k1%')
        #ax.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))
        ax.set_ylabel('Доля наблюдений')
        ax2.set_ylabel('k')
        plt.title(plt_title)
        plt.legend(('$\widetilde{k}$','Граница доверительного интервала 5%','Граница доверительного интервала 1%'),fontsize = 'small', loc = 1, frameon = False)
        plt.grid()

        if png_title:
            plt.savefig(png_title)
        #plt.show()
        
        MS1.rename(columns={'Новый сигнал по алгоритму ЦБ': 'Сигнал',
                           'Наблюдаемый уровень дефолта': 'DR',                  
                           'Наблюдаемый уровень дефолта + преобразование': 'DR + преобразование'}, inplace=True)
        plt.close(fig)
        return MS1[cols_to_return], fig



def bin_graf_log(df0, market=''):
    """ Строит график биномиального теста в логарифмической шкале
         на вход подается датафрейм из 7 столбцов. Данные обычные - нелогарифмированные!!!
         программа их преобразовывает и возвращает график.
    """
    df_rang = df0.copy()
    m0, ind, pd0, c0, c2, signal_before, signal_after, l_err, u_err, l_err99, u_err99 = df_rang.columns
    df_rang[c2] = df_rang[c2].apply(lambda x: 0 if x <= 0 else np.log(x / 0.0003))

    df_rang[c0] = df_rang[c0].apply(lambda x: 0 if x <= 0 else np.log(x / 0.0003))
    df_rang[pd0] = df_rang[pd0].apply(lambda x: 0 if x <= 0 else np.log(x / 0.0003))
    df_rang[l_err] = df_rang[l_err].apply(lambda x: 0 if x <= 0 else np.log(x / 0.0003))
    df_rang[u_err] = df_rang[u_err].apply(lambda x: 0 if x <= 0 else np.log(x / 0.0003))
    df_rang[l_err99] = df_rang[l_err99].apply(lambda x: 0 if x <= 0 else np.log(x / 0.0003))
    df_rang[u_err99] = df_rang[u_err99].apply(lambda x: 0 if x <= 0 else np.log(x / 0.0003))


    fig, ax = plt.subplots(1, 1, figsize=(10, 7))

    ax2 = ax.twinx()
    ax.bar(df_rang.index - 0.4, df_rang[ind] / df_rang[ind].sum() * 100, color='grey')

    ax.grid(None)
    ax2.grid(None)

    ax2.plot(df_rang.index, (df_rang[l_err99]), color='r', label='')
    ax2.plot(df_rang.index, (df_rang[u_err99]), color='r', label='границы 99% Дов. Инт.')

    ax2.scatter(df_rang.index, (df_rang[c0]), color='blue', label='Наблюдаемые дефолты')
    # if df_rang.columns.shape[0] == 7:
    ax2.scatter(df_rang.index, (df_rang[c2]), color='orange', label='Байесовский сдвиг на величину ЦТ')

    ax.set_title('Наблюдаемый и предсказываемый уровень дефолтов на фоне распределения по бакетам МШ.\n '
                 + market + '. Логарифмическая шкала PD.')
    ax.set_xticks(df_rang.index)
    ax.set_xticklabels(df_rang[m0], rotation='vertical')

    ax.set_ylabel('% наблюдений')
    ax2.set_ylabel('PD, %')

    ax2.hlines(-np.log(0.03), -0.5, df_rang.shape[0], color='white', linestyles='dashed')
    ax2.hlines(np.log(10 / 0.03), -0.5, df_rang.shape[0], color='white', linestyles='dashed')
    ax2.annotate('           1', xy=(df_rang.shape[0], -np.log(0.03)))
    ax2.annotate('          10', xy=(df_rang.shape[0], np.log(10 / 0.03)))

    ax2.set_yticklabels([0 if x < 0 else '' for x in ax2.get_yticks()])

    ax2.set_ylim([0, df_rang.c2.max() * 1.2])
    ax2.legend(loc=2)
    plt.close(fig)
    return fig

class Ks_test:
    """
    Тест Колмогорова-Смирнова
    ------------------------------------------
    Вход
    df: датафрейм
    s0: скор модели
    cel: целевая переменная
    border_list: массив с границами теста. По умолчанию [0.4, 0.3]

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: список с графиками
    """
    def __init__(self, df, border_list=None, **params): #s0, cel, 
        self.s0 = params['s0']
        self.cel = params['cel']
        
        self.df = df
#         self.s0 = s0
#         self.cel = cel
        self.tname = 'Тест Колмогорова-Смирнова'
        self.border_list = [0.4, 0.3] if border_list == None else border_list
        self.tbls = []
        self.grafls = []
        selfbig_plot, self.res = self.make_graf()
        self.tbls.append(self.make_table())
        self.grafls.append(selfbig_plot)

    def make_graf(self):
        cel = self.cel
        s0 = self.s0
        df = self.df[[cel, s0]].copy()
        df[s0]=np.abs(df[s0])

        df[s0] = df[s0] - np.min(df[s0])
        dfg = df[[s0, cel]].groupby(s0).agg({cel: ['count', 'sum']}).reset_index()
        dfg.columns = ['sc', 'cont', 'sm']
        dfg = dfg.sort_values('sc')

        dfg['cont'] = dfg['cont'] - dfg['sm']
        dfg['non_def'] = dfg['cont'].cumsum()
        dfg['yes_def'] = dfg['sm'].cumsum()
        ndmax = dfg['non_def'].max()
        ydmax = dfg['yes_def'].max()
        dfg['non_def'] = dfg['non_def'] / ndmax
        dfg['yes_def'] = dfg['yes_def'] / ydmax
        dfg['ks'] =  dfg['yes_def'] - dfg['non_def']
        ks_res = dfg['ks'].max()
        m = dfg['ks'].values.argmax()
        fig, ax = plt.subplots()

        ax.plot(dfg['sc'], dfg['yes_def'], color=[241 / 255, 204 / 255, 86 / 255], label='1')
        ax.plot(dfg['sc'], dfg['non_def'], color=[0, 170 / 255, 255 / 255], label='0')

        ax.set_title('Тест Колмогорова-Смирнова')

        ax.legend(loc='lower right')
        ax.vlines(dfg['sc'][m], dfg['yes_def'][m], dfg['non_def'][m], color=[234 / 255, 107 / 255, 80 / 255], lw=2)
        ax.annotate(str(round(ks_res, 4) * 100) + '%', xy=(dfg['sc'][m] * 1.1, dfg['non_def'][m]))
        ax.set_yticks(np.arange(0, 1.2, 0.1))
        ax.set_xlabel('Скор-балл модели')
        ax.set_ylabel('Вероятность')
        plt.close(fig)
        return fig, round(ks_res, 4)

    def make_table(self):
        df = self.df
        df_out = pd.DataFrame(columns=['Название теста', 'Результат',
                                       'Граница желтого сигнала',
                                       'Граница красного сигнала', 'Сигнал'])
        df_out.loc[1, :] = [self.tname, self.res, self.border_list[0], self.border_list[1], 'З'
        if self.res > self.border_list[0] else 'К' if self.res < self.border_list[1] else 'Ж']

        for i in range(1, 4):
            cl = df_out.columns[i]
            df_out[cl] = df_out[cl].apply(pd.to_numeric)

        return df_out


class Rel_gini_test:
    """
    Расчет относительного Gini
    ------------------------------------------
    Вход
    df1: первая выборка
    df2: вторая выборка
    col1: переменная из первой выборки
    col2: переменная из второй выборки
    cel1: целевая переменная из первой выборки
    cel2: целевая переменная из второй выборки
    name1: название первой выборки
    name2: название второй выборки
    alias: список псевдонимов переменных
    short_alias: список имен переменных для графиков
    border_list: список с границами желтого и красного сигналов для каждой переменной (по умолчанию 10%, 15%),.
    Пример задания списка - [[0.15, 0.1] for i in col]
    trust_level: доверительный интервал. По умолчанию 95%
    unidir: флаг, что монотонность скор балла и вероятности дефолта имеют
    одинаковые направления (по умолчанию True, т.е. направления совпадают)

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: список с граффиками
    """
    def __init__(self, df1, df2, name1='валидация', name2='разработка',
                 alias=None, short_alias=None, border_list=None, unidir=False, **params): #, col1, col2, cel1, cel2,
        col1 = params['col']
        col2 = params['col']
        cel1 = params['cel']
        cel2 = params['cel']

#         self.first_col = Gini_test(df1, col1, cel1, alias, short_alias, border_list, unidir=unidir)
#         self.second_col = Gini_test(df2, col2, cel2, alias, short_alias, border_list, unidir=unidir)
    
        self.first_col = Gini_test(df1, alias, short_alias, border_list, unidir=unidir, **params)
        self.second_col = Gini_test(df2, alias, short_alias, border_list, unidir=unidir, **params)
        
        self.tname = 'Относительный ' + self.first_col.tname

        if (alias==None) & (short_alias!=None):
            self.short_alias = short_alias
            self.alias = short_alias
        elif (alias!=None) & (short_alias==None):
            self.alias = alias
            self.short_alias = alias

        if (alias == None) & (short_alias == None):
            self.alias = col1
            self.short_alias=col1

        self.border_list = border_list
        if (border_list == None):
            self.border_list = [[0.05, 0.1] for i in col1]
        self.name1 = name1
        self.name2 = name2

        self.tbls = []
        self.grafls = []
        self.tbls.append(self.make_table())
#         self.grafls.append(self.make_big_plot())

    def make_table(self):
        df_out = pd.DataFrame() 
#         columns=['Фактор /модуль',
#                                        self.first_col.tname + ' на выборке ' + self.name1,
#                                        self.first_col.tname + ' на выборке ' + self.name2,
#                                        'Относительное изменение',
#                                        'Граница желтого сигнала',
#                                        'Граница красного сигнала',
#                                        'Сигнал'])
        
        df1 = self.first_col.tbls[0].iloc[-1:]
        df2 = self.second_col.tbls[0].iloc[-1:]
        
        df_out = df1
        
#         df_out[df_out.columns[0]] = ['model']
#         df_out[df_out.columns[1]] = [df1.iloc[0][df1.columns[1]]]
#         df_out[df_out.columns[2]] = [df2.iloc[0][df1.columns[1]]]
        df_out[self.first_col.tname + ' на выборке ' + self.name2] = [df2.iloc[0]['Джини']]
        df_out['Относительное изменение'] = (df_out['Джини'] - df_out[self.first_col.tname + ' на выборке ' + self.name2]) / df_out[self.first_col.tname + ' на выборке ' + self.name2]
#         df_out[df_out.columns[4]] = [x[0] for x in self.border_list[:1]]
#         df_out[df_out.columns[5]] = [x[1] for x in self.border_list[:1]]
        
        
#         df1 = self.first_col.tbls[0]
#         df2 = self.second_col.tbls[0]
        
#         df_out[df_out.columns[0]] = self.first_col.short_alias
#         df_out[df_out.columns[1]] = df1[df1.columns[1]]
#         df_out[df_out.columns[2]] = df2[df1.columns[1]]
#         df_out[df_out.columns[3]] = (df1[df1.columns[1]] - df2[df1.columns[1]]) / df2[df1.columns[1]]
#         df_out[df_out.columns[4]] = [x[0] for x in self.border_list]
#         df_out[df_out.columns[5]] = [x[1] for x in self.border_list]
        df_out['Относительное изменение сигнал'] = df_out.apply(
            lambda x: 'З' if np.abs(x['Относительное изменение']) < self.border_list[-1][0] else
            'К' if np.abs(x['Относительное изменение']) > self.border_list[-1][1] else 'Ж', axis=1)

        return df_out

    def make_big_plot(self):
        fout = self.tbls[0]

        fig, ax = plt.subplots(1, 1, figsize=(9, 6))
        width = 0.35

        df1 = self.first_col.tbls[0]
        df2 = self.second_col.tbls[0]
        err1 = (df1[df1.columns[3]] - df1[df1.columns[2]]) / 2
        err2 = (df2[df2.columns[3]] - df2[df2.columns[2]]) / 2

        ax.bar(fout.index - width / 2, fout[fout.columns[1]] * 100, yerr=(err1) * 100,
               align='center', width=width, color='orange', label=self.name1)

        ax.bar(fout.index + width / 2, fout[fout.columns[2]] * 100, yerr=(err2) * 100, align='center',
               width=width, color=[0, 170 / 255, 255 / 255], label=self.name2)

        bv_max = int(round(fout[fout.columns[1]].max() * 10 + 0.5) * 10 + 5)
        bv_min = int(round(fout[fout.columns[1]].min() * 10 - 0.5) * 10 - 5)
        bd_max = int(round(fout[fout.columns[2]].max() * 10 + 0.5) * 10 + 5)
        bd_min = int(round(fout[fout.columns[2]].min() * 10 - 0.5) * 10 - 5)

        b_max = np.max([bv_max, bd_max])
        b_min = np.min([bv_min, bd_min])

        ax.set_ylabel('%%')
        ax.set_title(self.tname + ' для ' + self.name1 + ' и ' + self.name2)

        ax.set_xticks(fout.index)
        ax.set_xticklabels(self.first_col.short_alias, rotation='vertical')
        nvars = fout.shape[0]
        red_sight = [e[0] * 100 for e in self.border_list]
        yellow_sight = [e[1] * 100 for e in self.border_list]
        ax.plot(range(-1, nvars + 1), [red_sight[0]] + red_sight + [red_sight[-1]],
                color=[241 / 255, 204 / 255, 86 / 255])
        ax.plot(range(-1, nvars + 1), [yellow_sight[0]] + yellow_sight + [yellow_sight[-1]],
                color=[234 / 255, 107 / 255, 80 / 255])
        ax.grid(axis='x')
        plt.yticks(range(b_min, b_max, 5))
        plt.close(fig)
        return fig


# недостаток - считает долго.
class Rel_mgini_test(Rel_gini_test):
    """
    Расчет относительного модифицированного Gini
    ------------------------------------------
    Вход
    df1: первая выборка
    df2: вторая выборка
    col1: переменная из первой выборки
    col2: переменная из второй выборки
    cel1: целевая переменная из первой выборки
    cel2: целевая переменная из второй выборки
    name1: название первой выборки
    name2: название второй выборки
    alias: список псевдонимов переменных
    short_alias: список имен переменных для графиков
    border_list: список с границами желтого и красного сигналов для каждой переменной (по умолчанию 1%0, 15%),.
    Пример задания списка - [[0.15, 0.1] for i in col]
    trust_level: доверительный интервал. По умолчанию 95%

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: список с граффиками
    """
    def __init__(self, df1, df2, col1, col2, cel1, cel2, name1='валидация', name2='разработка',
                 alias=None, short_alias=None, border_list=None):
        Rel_gini_test.__init__(self, df1, df2, col1, col2, cel1, cel2, name1=name1, name2=name2,
                 alias=alias, short_alias=short_alias, border_list=border_list)


        if (border_list == None):
            border_list = [[0.15, 0.1] for i in col1]
        self.first_col = Mgini_test(df1, col1, cel1, alias, short_alias, border_list)
        self.second_col = Mgini_test(df2, col2, cel2, alias, short_alias, border_list)
        self.tname = 'Сравнение модиф. Джини'
        self.border_list = border_list
        self.name1 = name1
        self.name2 = name2
        self.tbls = []
        self.grafls = []
        self.tbls.append(Rel_gini_test.make_table(self))
        self.grafls.append(Rel_gini_test.make_big_plot(self))

class Psi_disr_test:
    def __init__(self, df1, df2, col2=None, m='',m2='', alias=None, border_list=None, **params):
                
        df_cur = params['df_cur']
        self.col = params['col']
#         col2 = params['col']
        m = params['m0']
        m2 = params['m0']
        
        if alias == None:
            alias = []
        if col2 == None:
            col2 = []
        if border_list == None:
            border_list = []

        self.tname = 'Распределение значений факторов' 

        if alias == []:
            self.alias = self.col
        else:
            self.alias = alias
            
        self.tbls = []
        
        sample_dict = {
            'Разработка': df2,
            'Валидация': df1,
            'Тeкущий портфель': df_cur
        }
        self.grafls = self.plot_factor_dist(sample_dict)

    def plot_factor_dist(self, sample_dict):
        plot_list = []
        frame={}
        for i, fact in enumerate(self.col):
            fig, ax = plt.subplots()
            
            mx=-1
            f=0
            frame[fact]=pd.DataFrame()
            for key, tmp in sample_dict.items():
                tmp[fact] = tmp[fact].apply(lambda x: pd.to_numeric(x))
                tmp[fact] = np.round(tmp[fact], 5)
                frame[fact][key]=(tmp[fact].value_counts()/tmp[fact].shape[0]*100)
                if (f==0)|(frame[fact][key].max()>mx):
                    mx=frame[fact][key].max()
                    f=1
                  
            frame[fact]=frame[fact].sort_index()
            if frame[fact].shape[0]==2:
                width =1.2/frame[fact].shape[0]
            elif frame[fact].shape[0]==3:
                width =2.4/frame[fact].shape[0]
            else:
                width =3.5/frame[fact].shape[0]
    #             width =10/frame[fact].shape[0]
    
            frame[fact].plot(kind='bar', width=width,figsize=(16,10), 
                                color=[color_red, color_yellow, color_salad, color_green], ax=ax)
            w=width 
            ax.set(xlim=(-w, frame[fact].shape[0]-(1-w)), ylim=(0, mx*1.1))
            x = frame[fact].index
            ax.set_xticklabels(list(np.round(x,4)), fontsize=16, rotation = 'horizontal')
            plt.yticks(fontsize=16)
            fmt = '%.0f%%' # Format you want the ticks, e.g. '40%'
            ax.yaxis.set_major_formatter(mtick.FormatStrFormatter(fmt))
            plt.legend(loc='upper right', framealpha=1, fontsize = 14)
            plt.grid(color='w')
            plt.title('{}'.format(self.alias[i]), fontsize=16)
#             plt.show() 
#             plt.close()
            
            plot_list.append(fig)
            plt.close(fig)
        return plot_list
    
    
class Psi_test:
    """
    Расчет индекса стабильности популяции (PSI).
    ----------------------------------------------------------
    Вход
    df1: первая выборка
    df2: вторая выборка
    col: список переменных из первой выборки
    col2: список переменных из второй выборки (если совпадают с col, можно не задавать)
    m: имя столбца с рангами МШ в первой выбрке
    m2: имя столбца с рангами МШ во второй выбрке (если совпадают с m, можно не задавать)
    name1: название первой выборки
    name2: название второй выборки
    alias: список псевдонимов для переменных
    border_list: список с желтыми и красными границами теста для каждой переменной (по умолчанию 10%, 25%)
        Пример задания - [[0.1, 0.25] for i in col]
    weight_list: список с весами факторов (если не указаны рассчитываются по выборкам: max_score - min_score)
        Пример задания - [weight for i in col]

    Параметры класса
    self.tbls[0]: таблица с результатами теста
                  в случае пропусков или несовпадения уникальных значений к сигналу ставится * и код ошибки
                  Код ошибки:
                    0 – ошибок нету
                    1 – уникальные значения выборок не совпадают
                    2 – есть пропуски в выборках
                    3 – уникальные значения выборок не совпадают и есть пропуски
    self.tbls[1]: таблица с итоговым результатом теста
                  в случае наличия кода ошибки у одного из факторов к итоговому сигналу ставится *
    self.grafls: список с граффиками
    """
    def __init__(self, df1, df20, col2=None, m='',m2='', name1='валидация', name2='разработка',
                 alias=None, border_list=None, weight_list=None, **params): #, col1, col2, cel1, cel2,
        col = params['col']
#         col2 = params['col']
        m = params['m0']
        m2 = params['m0']
        
        if alias == None:
            alias = []
        if col2 == None:
            col2 = []
        if border_list == None:
            border_list = []

        self.tname = 'Индекс PSI ' + name1 + '-' + name2
        self.border_list = border_list
        if (border_list == []):
            self.border_list = [[0.1, 0.25] for i in col]
        self.name1 = name1
        self.name2 = name2
        self.df1 = df1
        if col2==[]:
            col2=col
        df2 = df20.copy()
        for i in range(len(col)):
            if col[i]!=col2[i]:
                df2.rename(columns={col2[i]:col[i]},inplace=True)
        if m!='':
            df2.rename(columns={m2: m}, inplace=True)

        self.df2 = df2
        self.col = col
        self.m = m

        if alias == []:
            self.alias = self.col
        else:
            self.alias = alias
            
        # Расчет весов
        if weight_list == None:
            self.weight_list = []
            for i in range(len(self.alias)):
                unique_val = set(self.df1[self.col[i]]).intersection(set(self.df2[self.col[i]]))
                self.weight_list.append(max(unique_val) - min(unique_val))
        else:
            self.weight_list = weight_list
        
        # Списки с результатами и графиками теста
        self.tbls = []
        self.grafls = []

        self.tbls.append(self.make_table())
        self.tbls.append(self.agg_psi())

        if m != '':
            self.grafls.append(self.make_big_plot())
        self.grafls+=self.make_list_plot()

    def make_table(self):
        df_out = pd.DataFrame(columns=['Фактор / Модуль / Модель', 'Вес', 'Population Stability Index (PSI)',
                                       'Пороговое значение Желтого', 'Пороговое значение Красного', 'Сигнал', 'Код ошибки'])
        
        assert(all(isinstance(x, (int, float)) for x in self.weight_list)), 'Веса должны быть числами' 
        weight_sum = sum(self.weight_list)
        
        for i in range(len(self.alias)):
            psi_result = psi(self.df1[self.col[i]], self.df2[self.col[i]])
            
            test_result = ""
            if psi_result[1] != 0:
                test_result = "*"
            if np.abs(psi_result[0]) < self.border_list[i][0]:
                test_result = "З" + test_result
            elif np.abs(psi_result[0]) > self.border_list[i][1]:
                test_result = "К" + test_result
            else:
                test_result = "Ж" + test_result
            
            df_out.loc[i, :7] = [self.alias[i],
                                 round(self.weight_list[i]/weight_sum, 2),
                                 psi_result[0],
                                 self.border_list[i][0], 
                                 self.border_list[i][1],
                                 test_result, 
                                 psi_result[1] if psi_result[1] else "-"]
            
        for i in range(1, 5):
            cl = df_out.columns[i]
            df_out[cl] = df_out[cl].apply(pd.to_numeric)

        return df_out

    def make_big_plot(self):
        dfm1 = (self.df1[self.m].value_counts() / self.df1.shape[0]).sort_index().reset_index()
        dfm2 = (self.df2[self.m].value_counts() / self.df2.shape[0]).sort_index().reset_index()
        df_chart = dfm1.merge(dfm2, on='index', how='outer').fillna(0)

        fig, ax = plt.subplots(1, 1, figsize=(10, 7))
        width = 0.35
        ax.bar(df_chart.index - width / 2, df_chart[df_chart.columns[1]], align='center',
               width=width, label=self.name1)
        ax.bar(df_chart.index + width / 2, df_chart[df_chart.columns[2]],
               align='center', width=width, color='red', label=self.name2)

        ax.set_ylabel('%%')
        ax.set_title('Распределение по рангам мастер шкалы.')

        ax.set_xticks(df_chart.index)
        ax.set_xticklabels(df_chart['index'], rotation='vertical')
        ax.legend()
        plt.close(fig)
        return fig

    def make_list_plot(self):
        ls = []
        for i in range(len(self.col)):
            a = self.col[i]

            dfm1 = (self.df1[a].value_counts() / self.df1.shape[0]).sort_index().reset_index()
            dfm2 = (self.df2[a].value_counts() / self.df2.shape[0]).sort_index().reset_index()
            df_chart = dfm1.merge(dfm2, on='index', how='outer').fillna(0)

            fig, ax = plt.subplots()
            width = 0.35
            ax.bar(df_chart.index - width / 2, df_chart[df_chart.columns[1]], align='center',
                   width=width, label=self.name1)
            ax.bar(df_chart.index + width / 2, df_chart[df_chart.columns[2]],
                   align='center', width=width, color='red', label=self.name2)

            ax.set_ylabel('%%')
            ax.set_title('Распределение ' + self.alias[i])
            ax.set_xticks(df_chart.index)
            ax.set_xticklabels(df_chart['index'], rotation='vertical')
            ax.legend()

            ls.append(fig)
            plt.close(fig)
        return ls
    
    def agg_psi(self):
        df_out = pd.DataFrame(columns=['Сигнал', 'Доля красных', 'Доля желтых'])
        all_sum = 0
        red_sum = 0
        yellow_sum = 0
        warning_flg = 0
        
        for i in range(len(self.alias)):
            all_sum += self.weight_list[i]
            if self.tbls[0].iloc[i, 5] in ["Ж", "Ж*"]:
                yellow_sum += self.weight_list[i]
            if self.tbls[0].iloc[i, 5] in ["К", "К*"]:
                red_sum += self.weight_list[i]
            if self.tbls[0].iloc[i, 5].find("*") != -1:
                warning_flg = 1
        
        signal = ""
        if warning_flg:
            signal += "*"
        
        if red_sum/all_sum > 0.25:
            signal = "К" + signal
        elif red_sum/all_sum > 0.1:
            signal = "Ж" + signal
        elif yellow_sum/all_sum > 0.25:
            signal = "Ж" + signal
        else:
            signal = "З" + signal
        df_out.loc[0, :] = [signal, red_sum/all_sum, yellow_sum/all_sum]
        
        return df_out

class LEK_test:
    def __init__(self, df, **params):
        self.tbls = []
        self.grafls = []

        self.tbls.append(self.make_table(df))
        print(self.tbls[0])

    def make_table(self, df):
        test = pd.DataFrame()
#         test['podsegment']=['Средние','Крупнейшие']
        test['Доля ЛК1 по подсегментам, %'] = [df[df['lec1']==1].shape[0]/df.shape[0]]
        test['Граница желтого сигнала, %'] = 5
        test['Граница красного сигнала, %'] = 7
        test['Сигнал']=np.where(test['Доля ЛК1 по подсегментам, %']>7,'К',np.where(test['Доля ЛК1 по подсегментам, %']>5,'Ж','З'))
#         test=test.set_index('podsegment') 
        return test
    

class R2_test:
    """
    Расчет R^2
    --------------------------------------
    df: датафрейм
    col: список переменных
    cel: целевая переменная
    alias: список псевдонимов

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    """
    def __init__(self, df, alias=None, border_list=None, **params): #col, cel
        col = params['col']
        cel = params['cel']
        
        if alias == None:
            alias = []

        if border_list == None:
            self.border_list = [0.4, 0.3]
        else:
            self.border_list = border_list

        self.tname = 'r2'
        self.col = col
        self.df = df
        self.cel = cel
        if alias == []:
            self.alias = self.col
        else:
            self.alias = alias

        res, self.tbl_plus = self.make_table()
        self.tbl = pd.DataFrame(columns=['Наименование теста',
                                         'Значение', 'Граница желтого сигнала',
                                         'Граница красного сигнала',
                                         'Сигнал'])
        self.tbl.loc[0, :] = ['R2', res, self.border_list[0], self.border_list[1],
                              'З' if res > self.border_list[0] else 'К'if res < self.border_list[1] else 'Ж']
        self.tbls = []
        self.grafls = []
        for i in range(1, 4):
            cl = self.tbl.columns[i]
            self.tbl[cl] = self.tbl[cl].apply(pd.to_numeric)

        self.tbls.append(self.tbl)
        self.tbls.append(self.tbl_plus)

    def make_table(self):
        ls = self.col
        ng = self.cel
        df = self.df
        reg1 = linear_model.LinearRegression()
        reg1.fit(df[ls], df[ng])
        res =  reg1.score(df[ls], df[ng])

        # считаем ошибку регрессии
        n = df.shape[0]
        k = len(ls)
        SER = (4 * res * (1 - res) ** 2 * (n - k - 1) ** 2 / (n ** 2 - 1) / (n + 3)) ** 0.5

        # считаем ошибку коэфф
        df_regr = pd.DataFrame(columns=['f_name', 'k', 'stats', 'p-val'])
        for i in range(len(ls)):
            w = ls[i]
            sr = df[w].mean()
            ts = reg1.coef_[i] / (df[w].std() ** 2) * SER * n
            df_regr.loc[df_regr.shape[0], :] = self.alias[i], reg1.coef_[i], ts, sts.t.sf(ts, n)
        df_regr.columns = ['Переменная', 'Коэфф', 'Ошибка', 'P-value']
        return res, df_regr


class Chi2_test:
    """
    Тест Хосмера-Лемешева
    ------------------------------------------
    Вход
    df: датафрейм
    odr: столбец с дефолтами
    pd: столбец с pd
    m: столбец с рейтингами МШ
    ct: значение CT
    sign_levels: уровни значимости (по умолчанию 5%, 1%)

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    """
    def __init__(self, df, odr, pd2, m, ct=1, sign_levels=None):
        self.tname = 'Хосмер-Лемешев'

        self.ct=ct
        self.tbls = []
        self.grafls = []
        n = df.shape[0]
        df_rang = df.groupby(m).agg({pd2: 'sum',odr: ['count', 'mean']}).reset_index()


        df_rang.columns = df_rang.columns.droplevel(0)
        df_rang.rename(columns = { '':'ms','sum': 'pd1' ,'mean': 'odr','count':'cnt'},inplace =True)

        df_rang = df_rang[['ms', 'cnt', 'odr', 'pd1']]

        if sign_levels == None:
            sign_levels = [0.05, 0.01]

        assert len(sign_levels) == 2, "len(sign_levels) != 2 "
        assert sign_levels[0] > sign_levels[1], "The values in sign_levels must be in ascending order"

        if self.ct != 1:
            dr = (df_rang['odr'] * df_rang['cnt']).sum() / df_rang['cnt'].sum()
            df_rang['odr'] = df_rang['odr'].apply(
                lambda x: x * self.ct / dr / (x * self.ct / dr + (1 - x) * (1 - self.ct) / (1 - dr)))
        else:
            df_rang['odr'] = df_rang['odr']
        # df_rang = df_rang[['ms', 'cnt', 'odr', 'pd1']]

        df_rang.odr = df_rang.odr
        df_rang.pd1 = df_rang.pd1 / df_rang.cnt
        df_rang['ti'] = df_rang.apply(lambda x: 0 if x.pd1==0 else round(((x.pd1 - x.odr) ** 2) * x.cnt / x.pd1 / (1 - x.pd1),2),
                                      axis=1)
        sm = df_rang.ti.sum()
        nh = df_rang.shape[0] - 2
        res = round(1 - sts.chi2.cdf(sm, nh), 2)

        tl_1 = 1 - sign_levels[0]
        tl_2 = 1 - sign_levels[1]

        top_1 = "Граница {}% интервала $\chi^2({})$".format(int(tl_1 * 100), nh)
        top_2 = "Граница {}% интервала $\chi^2({})$".format(int(tl_2 * 100), nh)

        z_1 = round(sts.chi2.ppf(tl_1, nh), 2)
        z_2 = round(sts.chi2.ppf(tl_2, nh), 2)

        df_rang.loc[df_rang.shape[0], df_rang.columns] = 'HL_statistic (∑ Ti)',0,0,0, sm
        df_rang.loc[df_rang.shape[0], df_rang.columns] = 'p-value',0,0,0, res
        # df_rang.loc[df_rang.shape[0], df_rang.columns] = r'$\chi^2$' + top_1 ,0,0,0, z_95
        # df_rang.loc[df_rang.shape[0], df_rang.columns] = 'Граница красной зоны:',0,0,0, z_99
        df_rang.loc[df_rang.shape[0], df_rang.columns] = top_1, 0, 0, 0, z_1
        df_rang.loc[df_rang.shape[0], df_rang.columns] = top_2, 0, 0, 0, z_2
        # df_rang.loc[df_rang.shape[0], df_rang.columns] = 'Сигнал',0,0,0, 'З' if res < top_green else 'Ж' if res < top_yellow else 'К'
        df_rang.loc[df_rang.shape[0], df_rang.columns] = 'Сигнал',0,0,0, 'З' if res > z_1 else 'К' if res < z_2 else 'Ж'

        df_rang['odr'] = df_rang['odr'].apply(pd.to_numeric)
        df_rang['pd1'] = df_rang['pd1'].apply(pd.to_numeric)
        df_rang['cnt'] = df_rang['cnt'].apply(lambda x:int(x))

        df_rang.columns = ['Разряд Шкалы', 'Количество наблюдений',
                           'Наблюдаемая частота дефолтов (скорр.)','Расчетное значение PD',
                           'Статистика Хи-квадрат для разряда рейтинговой шкалы (Ti)']


        self.tbls.append(df_rang)



class Cal_test:
    """
    Вход: датафрейм, флаг дефолта, PD  модели
    """
    def __init__(self, df, ct=1, bayes_check=False, **params):
        pd1 = params['pd0']
        cel = params['cel']
        self.tname = 'Калибровка'
        self.tbls = []
        self.grafls = []
        df_rang = pd.DataFrame()#columns=['Среднее PD модели',
#                                         'Среднее DR, рассчитанное на основе фактических данных',
#                                         'Количество наблюдений',
#                                         '1%','5%', '95%','99%', 'Сигнал'])
        ci_95 = sts.norm.interval(0.95)[1]
        ci_99 = sts.norm.interval(0.99)[1]
        n = df.shape[0]
        pd1_v = df[pd1].sum()/n
        df_rang['Среднее PD модели'] = [pd1_v]
        
#         odr_v=df[odr].sum()/n

#         dr = df[cel].mean()
#         bayes_flg = False if np.round(dr, 4) == ct else True
        #print('Bayes check:', bayes_check, ct)#, 'flg:', bayes_flg)
        if bayes_check:# and bayes_flg:
            res_tbl = Bin_test_kpk(df, ct=ct, **params).tbls[0]
            cel_v = (res_tbl['DR + преобразование'] * res_tbl['Количество наблюдений']).sum()/res_tbl['Количество наблюдений'].sum()
            df_rang['DR с трансформацией Байеса'] = [cel_v]
            
        else:
            cel_v=df[cel].sum()/n
            df_rang['Среднее DR, рассчитанное на основе фактических данных'] = [cel_v]
            
        df_rang['Количество наблюдений'] = [n]
            
        k = (np.abs(pd1_v * (1 - pd1_v) / n)) ** 0.5
        df_rang.loc[0, ['1%','5%', '95%', '99%', 'Сигнал']] = pd1_v-k*ci_99,pd1_v-k*ci_95,\
                         pd1_v+k*ci_95,pd1_v+k*ci_99,\
                        'З' if np.abs(cel_v-pd1_v)<k*ci_95 else \
                        'К' if cel_v-pd1_v>k*ci_99 else 'Ж'
        
#         df_rang.loc[1,:]=pd1_v,cel_v,n,pd1_v-k*ci_99,pd1_v-k*ci_95,\
#                          pd1_v+k*ci_95,pd1_v+k*ci_99,\
#                         'З' if np.abs(cel_v-pd1_v)<k*ci_95 else \
#                         'К' if np.abs(cel_v-pd1_v)>k*ci_99 else 'Ж'
        for i in range(0, 7):
            cl = df_rang.columns[i]
            df_rang[cl] = df_rang[cl].apply(pd.to_numeric)

        self.tbls.append(df_rang)


class Herfindal_test:
    """
    Расчет теста Херфиндаля
    -----------------------
    Вход
    df: датафрейм
    m0: столбец с рейтингами МШ
    border_list: список с границами теста (по умолчанию Ж - 10%, К - 20%)
    j: общее количество разрядов рейтинговой шкалы (за исключением дефолтного разряда). Если j=-1, 
    то значение равно количеству грейдов в таблице.

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    """
    def __init__(self, df, border_list=None, j=-1, **params): #m0
        m0 = params['m0']
        
        self.tname = 'Херфиндаль'
        self.tbls = []
        self.grafls = []
        if border_list == None:
            border_list = [0.1, 0.2]
        
        check = lambda x: 'З' if x <= border_list[0] \
                      else 'К' if x > border_list[1] \
                      else 'Ж' 

        df_rang = pd.DataFrame(columns=['Наименование теста',
                                        'Значение', 
                                        'Значение скорр.',
                                        'Граница желтого сигнала',
                                        'Граница красного сигнала',
                                        'Сигнал',
                                        'Сигнал (после корр.)'])
        
        
        kherr = (df[m0].value_counts() / df.shape[0]) ** 2
        j = kherr.shape[0] if j==-1 else j
        kherr = kherr.sum()
        kherr_corr = (kherr - 1/j)/(1 - 1/j)
        df_rang.loc[1,:] = 'Индекс Херфиндаля' , kherr, kherr_corr, \
                            border_list[0], border_list[1], \
                            check(kherr), check(kherr_corr)        
        for i in range(1, 5):
            cl = df_rang.columns[i]
            df_rang[cl] = df_rang[cl].apply(pd.to_numeric)

        self.tbls.append(df_rang)
        
        func_light_perc = lambda value: '{:.0%}'.format(value)
        func_value_perc = lambda value: '{:.2%}'.format(value)
        self.kpk_tbl = pd.DataFrame([[func_value_perc(kherr_corr), 
                                     f'>{func_light_perc(border_list[1])}', 
                                     f'({func_light_perc(border_list[0])}; {func_light_perc(border_list[1])})',
                                     f'<= {func_light_perc(border_list[0])}', 
                                     light_rating_dict[check(kherr_corr)]]],
                                    columns=['value', 'threshold_red', 'threshold_yellow', 'threshold_green', 'rating']
                                    )



# Диагональный тест. Вход: датафрейм, цифры Рангов модель, цифры рангов Учитель.
class Diag_test:
    """
    Вход: df = датафрейм, c0 = цифровые номера рангов - предсказания модели
    r0 = цифровые номера рангов учителя
    """
    def __init__(self, df, c0, r0):
        self.tname = 'Диагональный тест'
        self.tbls = []
        self.grafls = []
        self.df = df
        self.c0 = c0
        self.r0 = r0
        if df[r0].unique().shape[0] > 2:
            self.make_test()

    def make_test(self):
        r0 = self.r0
        c0 = self.c0
        df = self.df

        tm = df[[r0, c0]]
        # вычисляем значения для анализа тренда, дисперсии, выбросов
        tm['res'] = tm[r0] - tm[c0]
        df_res = pd.DataFrame(columns=['Без Объединения', 'В пределах 2-х нотчей', 'В пределах 4-х нотчей'])
        df_res.loc[0, :] = [0, 0, 0]
        df_res.loc[1, :] = [0, 0, 0]
        df_res.loc[2, :] = [0, 0, 0]
        df_res.index = ['над диаг', 'на диаг', 'под диаг']

        df_res.iloc[0, 0] = tm[tm.res > 0].shape[0]
        df_res.iloc[1, 0] = tm[tm.res == 0].shape[0]
        df_res.iloc[2, 0] = tm[tm.res < 0].shape[0]

        df_res.iloc[0, 1] = tm[tm.res > 2].shape[0]
        df_res.iloc[1, 1] = tm[(tm.res <= 2) & ((tm.res >= -2))].shape[0]
        df_res.iloc[2, 1] = tm[tm.res < -2].shape[0]

        df_res.iloc[0, 2] = tm[tm.res > 4].shape[0]
        df_res.iloc[1, 2] = tm[(tm.res <= 4) & ((tm.res >= -4))].shape[0]
        df_res.iloc[2, 2] = tm[tm.res < -4].shape[0]

        tm = df[[r0, c0]]

        for i in range(21):
            tm.loc[tm.shape[0], :] = [0, i]
        for j in range(21):
            tm.loc[tm.shape[0], :] = [j, 0]

        ttn = pd.crosstab(tm[r0], tm[c0], margins=False)
        ttn = ttn.drop(0, axis=1).drop(0, axis=0)
        ttn = ttn.where(ttn != 0, other='')

        spr_ms = ['MA1', 'MA2', 'MA3', 'MB1', 'MB2', 'MB3', 'MC1', 'MC2', 'MC3', 'MD1', 'MD2',
                  'MD3', 'ME1', 'ME2', 'ME3', 'MF1', 'MF2', 'MF3', 'MG1', 'MG2', 'MG3']

        #for j in range(ttn.shape[1]):
        #    ttn.columns[j] = spr_ms[j]
        #for i in range(ttn.shape[0]):
        #    ttn.index[i] = spr_ms[i]

        df_out = pd.DataFrame(columns=['Название теста', 'Результат', 'Граница желтого сигнала',
                                       'Граница красного сигнала'])
        df_out.loc[0, :] = 'Тренд', np.abs(df_res.iloc[0, 0] - df_res.iloc[2, 0]) / tm.shape[0], 0.25, 0.5
        df_out.loc[0, :] = 'Дисперсия', (df_res.iloc[0, 1] + df_res.iloc[2, 1]) / tm.shape[0], 0.15, 0.25
        df_out.loc[0, :] = 'Выбросы', (df_res.iloc[0, 2] + df_res.iloc[2, 2]) / tm.shape[0], 0.05, 0.1
        df_out['Сигнал'] = df_out.apply(lambda x: 'З' if x[df_out.columns[1]] < x[df_out.columns[2]] else
        'К' if x[df_out.columns[1]] > x[df_out.columns[3]] else 'Ж', axis=1)

        self.tbls.append(df_out)
        self.tbls.append(ttn)


class Corr_test:
    """
    Рассчитывает коэффициенты корреляции Пирсона и Спирмена
    -------------------------------------------
    Вход
    df: датафрейм
    col: столбцы с переменными
    alias: псевдонимы
    short_alias: короткие псевдонимы
    border_list: границы теста (по умолчанию 60% и 75%)

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: матрицы корреляций
    """
    def __init__(self, df, alias=None, short_alias=None, border_list=None, **params): #col
        col = params['col']
        
        if (alias==None) & (short_alias!=None):
            self.short_alias = short_alias
            self.alias = short_alias
        elif (alias!=None) & (short_alias==None):
            self.alias = alias
            self.short_alias = alias

        if (alias == None) & (short_alias == None):
            self.alias = col
            self.short_alias=col

        if border_list == None:
            border_list = [0.6, 0.75]


        self.tname = 'Корреляция'
        self.tbls = []
        self.tbls.append(['00'])
        self.grafls = []

        df_res = pd.DataFrame(columns=['Метод',
                                       'Максимальное Значение',
                                       'Граница желтого сигнала',
                                       'Граница красного сигнала',
                                       'Сигнал'])
        df_corr1 = round(df[col].corr(), 2)
        df_corr2 = round(df[col].corr(method='spearman'), 2)

        df_corr1.columns = self.short_alias
        df_corr1.index = self.short_alias
        fig, ax = plt.subplots()
        fig.clf()
        fig = sns.heatmap(df_corr1, annot=True).get_figure()
        fig.suptitle('Корреляция Пирсона')
        self.grafls.append(fig)

        df_corr1.columns = self.alias
        df_corr1.index = self.alias

        self.tbls.append(df_corr1)

        df_corr2.columns = self.short_alias
        df_corr2.index = self.short_alias
        fig, ax = plt.subplots()
        fig.clf()
        fig = sns.heatmap(df_corr2, annot=True).get_figure()
        fig.suptitle('Корреляция Спирмана')
        self.grafls.append(fig)
        df_corr2.columns = self.alias
        df_corr2.index = self.alias
        self.tbls.append(df_corr2)

        for i in range(df_corr1.shape[0]):
            df_corr1.iloc[i, i] = 0
            df_corr2.iloc[i, i] = 0

        k1 = df_corr1.max().max()
        df_res.loc[0, :] = 'Корреляция Пирсона', k1, border_list[0], border_list[1], \
                           'З' if k1 < border_list[0] else 'К' if k1 > border_list[1] else 'Ж'

        k2 = df_corr2.max().max()
        df_res.loc[1, :] = 'Корреляция Спирмана', k2, border_list[0], border_list[1], \
                           'З' if k2 < border_list[0] else 'К' if k2 > border_list[1] else 'Ж'

        for i in range(1, 4):
            cl = df_res.columns[i]
            df_res[cl] = df_res[cl].apply(pd.to_numeric)
        self.tbls[0] = df_res
        plt.close(fig)


class Vif_test:
    """
    Рассчитывает коэффициенты VIF
    -------------------------------------------
    Вход
    df: датафрейм
    col: столбцы с переменными
    s: название переменной скор-балла (для удаления из списка)
    alias: псевдонимы
    short_alias: короткие псевдонимы
    border_list: границы теста (по умолчанию 4% и 10%)

    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: матрицы корреляций
    """

    def __init__(self, df, alias=None, short_alias=None, border_list=None, **params): #col, s
        col = params['col']
        s = params['s0']

        if (alias==None) & (short_alias!=None):
            self.short_alias = short_alias
            self.alias = short_alias
        elif (alias!=None) & (short_alias==None):
            self.alias = alias
            self.short_alias = alias

        if (alias == None) & (short_alias == None):
            self.alias = col
            self.short_alias=col

        if border_list == None:
            border_list = [0.04, 0.1]



        self.tname = 'VIF'
        self.tbls = []
        self.grafls = []

        vif = pd.DataFrame(columns=['Var', 'VIF', 'yell', 're'])

        col2=col.copy()
        if s in col2:
            col2.remove(s)
        for i in range(len(col2)):
            ls = col2.copy()
            f = ls[i - 1]
            ls.remove(f)

            reg1 = linear_model.LinearRegression()
            reg1.fit(df[ls], df[f])
            vif.loc[i, :] = [self.alias[i],
                             # 1 / (1 - metrics.r2_score(df[f], reg1.predict(df[ls]))),
                             0.01 / (1 - reg1.score(df[ls], df[f])),
                             border_list[0], border_list[1]]

        fig, ax = plt.subplots(1, 1, figsize=(10, 7))

        ax.bar(vif.index, vif.VIF*100, align='center')
        ax.set_xticks(vif.index)
        ax.set_xticklabels(vif.Var, rotation='vertical')
        ax.set_ylabel('%')

        ax.plot(range(-1, vif.shape[0] + 1), [4 for e in range(vif.shape[0] + 2)], color='yellow')
        ax.plot(range(-1, vif.shape[0] + 1), [10 for e in range(vif.shape[0] + 2)], color='red')
        ax.set_title('Фактор инфляции дисперсии')
        vif['Сигнал'] = vif.apply(lambda x: 'З' if x[vif.columns[1]] < x[vif.columns[2]] else
        'К' if x[vif.columns[1]] > x[vif.columns[3]] else 'Ж', axis=1)
        vif.columns = ['Переменная', 'Значение VIF', 'Граница желтого сигнала',
                       'Граница красного сигнала', 'Сигнал']
        for i in range(1, 4):
            cl = vif.columns[i]
            vif[cl] = vif[cl].apply(pd.to_numeric)


        self.tbls.append(vif)
        self.grafls.append(fig)
        plt.close(fig)
        
        
        
class Woe_test(Iv_test):
    """
    Параметры класса
    self.tbls[0]: таблица с результатами теста
    self.grafls: список с графиками WOE
    """
    def __init__(self, df, alias=None, short_alias=None, border_list=None, trust_level=0.95, unidir=True,
                 sample_name='', df_ms=None, ms_col='', ms_col1='', ms_pd_col='', model_type = None, **params): #col, cel
        col = params['col']
        self.factor_list = col
        cel = params['cel']
        
        self.model_type = model_type
        
        if df_ms is not None:
            alias_, short_alias_ = None, None
            df_tmp = df.merge(df_ms, left_on=ms_col, right_on=ms_col1)
            df_tmp[ms_pd_col] = -df_tmp[ms_pd_col]
            col_ = col + [ms_pd_col]
            if alias is not None:
                alias_ = alias + ['model']
            if short_alias is not None:
                short_alias_ = short_alias + ['model']
        else:
            df_tmp, alias_, short_alias_, col_ = df, alias, short_alias, col,

        Iv_test.__init__(self, df_tmp, alias=alias_, short_alias=short_alias_,
                         border_list=border_list, trust_level=trust_level, sample_name=sample_name, 
                         **{'col': col_, 'cel': cel}) #, col_, cel
        
        self.sample_name = sample_name
        self.unidir = unidir
            
        self.grafls = []
        self.tbls = []
        
        self.make_woe_list()
        

    def make_woe_list(self):
        df = self.df
        cel = self.cel

        res_table = pd.DataFrame(columns=['Фактор', 'Монотонность', 'Разница в соседних бакетах', 
                                          'Низкая концентрация наблюдений в бакете',
                                          'Высокая концентрация наблюдений в бакете'])

        for vv in self.factor_list: #[:-1]:
            if df[vv].unique().shape[0] < 20:
                woe_plot, res_dict = self.make_woe_quant(vv)
            else:
                print('contin:', vv)
                woe_plot, res_dict = self.make_woe_contin(vv)
                
            self.grafls.append(woe_plot)
            res_table = res_table.append(res_dict, ignore_index=True)
            
        self.tbls.append(res_table.replace({True: '+', False: '–'}))

    def make_woe_quant(self, vv):
#         reverse_param = self.df[]
        
        f_name = self.short_alias[self.col.index(vv)]
        cel = self.cel
        ddd = self.df[[vv, cel]].groupby(vv).agg({vv: 'count', cel: 'sum'})
        ddd.index.name = 'i1'
        ddd.reset_index(inplace=True)

        ddd['pd'] = ddd[cel] / ddd[vv]

        fig, ax = plt.subplots()
        ax2 = ax.twinx()
        s_norm = np.sum(ddd[vv]) / 100

        ax.bar(ddd.index, ddd[vv] / s_norm, align='center', width=0.5)
        ax2.plot(ddd.index, ddd.pd * 100, color='orange')
        ax.set_title(self.sample_name + '\n' + f_name)
        for r in ddd.iterrows():
            ax2.annotate(round(r[1].pd * 100, 2), xy=(r[0], r[1]['pd'] * 100))
        ax2.grid(None)
        ax.set_ylabel('% наблюдений')
        if self.model_type == 'LGD':
            ax2.set_ylabel('Cured, %')
        else:
            ax2.set_ylabel('DR, %')
            
        ax2.set_ylim([0, ddd.pd.max() * 130])

        ax.set_xticks(ddd.index)
        ddd.i1 = ddd.i1.apply(lambda x: pd.to_numeric(x))
        ax.set_xticklabels(ddd.i1.round(4))
        
        res_dict = ({'Фактор': f_name, 'Монотонность': self.check_monotonic(ddd.pd), 
                     'Разница в соседних бакетах': self.check_rel_pd(ddd.pd), 
                     'Низкая концентрация наблюдений в бакете': self.check_small_size_bucket(ddd[vv] / s_norm / 100),
                     'Высокая концентрация наблюдений в бакете': self.check_large_size_bucket(ddd[vv] / s_norm / 100)
                    })
        plt.close(fig)
        return (fig, res_dict)
    
    def check_small_size_bucket(self, size_bucket, small_size_threshold=0.03):
        return all(size_bucket > small_size_threshold)
    
    def check_large_size_bucket(self, size_bucket, large_size_threshold=0.7):
        return all(size_bucket < large_size_threshold)
    
    def check_rel_pd(self, pd_list, rel_threshold=0.03):
        for i, _ in enumerate(pd_list[:-1]):
            if abs((pd_list[i+1] - pd_list[i]) / pd_list[i+1]) < rel_threshold:
                return False
        return True
            
    def check_monotonic(self, arr, reverse=False):
#         if reverse:
#             return all(arr[i] >= arr[i+1] for i in range(len(arr)-1))
        return all(arr[i] <= arr[i+1] for i in range(len(arr)-1)) or all(arr[i] >= arr[i+1] for i in range(len(arr)-1))

    def make_woe_contin(self, vv, n=5):
        df = self.df
        cel = self.cel
        f_name = self.short_alias[self.col.index(vv)]

        ddd0 = df[[vv, cel]].copy()
        ddd0[vv] = np.abs(ddd0[vv])
        dmax = ddd0[vv].max()
        dmin = ddd0[vv].min()

        ddd0['norm'] = (ddd0[vv] - dmin) / (dmax - dmin) * 100
        ddd0['n_baket'] = (ddd0['norm'] // n) / 100

        ddd1 = ddd0.groupby('n_baket').agg({vv: 'count', 'norm': 'mean', cel: 'mean'}).reset_index()
        ddd1 = ddd1.sort_values('norm')

        fig, ax = plt.subplots()
        ax2 = ax.twinx()

        ax.set_ylabel(f_name + ' , %')
        ax.plot(ddd1.index, ddd1.norm, color='green', label=f_name)
        
        ax2.grid(None)
        ax2.plot(ddd1.index, ddd1[cel], color='red')#, label='DR')
        if self.model_type == 'LGD':
            ax2.set_ylabel('Cured, %')
        else:
            ax2.set_ylabel(' DR, % ')
            
        ax.bar(ddd1.index, ddd1[vv] / ddd0.shape[0] * 100, align='center')
        ax.legend(loc=2)
        ax2.legend(loc=1)

        ax.set_xticklabels([])
        ax.set_title(self.sample_name + '\n' + f_name + ' и PD на фоне распределения.\n Бакеты по ' + str(n) + '%')
        
        res_dict = ({'Фактор': f_name, 'Монотонность': self.check_monotonic(ddd1[cel]), 
             'Разница в соседних бакетах': self.check_rel_pd(ddd1[cel]), 
             'Низкая концентрация наблюдений в бакете': self.check_small_size_bucket(ddd1[vv] / ddd0.shape[0] * 100),
             'Высокая концентрация наблюдений в бакете': self.check_large_size_bucket(ddd1[vv] / ddd0.shape[0] * 100)
            })
        plt.close(fig)
        return (fig, res_dict)



#Функции модуля. Вывод.


# выводит таблицу с результатом теста  в эксель в  формате результатов теста (светофор)
def tbl_res_out(df, ws):
    """На входе: датафрейм для публикации, лист excel"""
    ck = 'ea6b50'
    cz = '78B497'
    cs = 'f1cc56'


    col_size = df.shape[1]

    itr = 0
    for col in df.columns:
        itr += 1
        hd = ws.cell(row=1, column=itr)
        hd.value = (col)
        hd.fill = PatternFill(start_color='002882', end_color='002882', fill_type="solid")
        hd.font = Font(bold=True, color='ffffff')
        #         print(col,itr,str(df[col].dtype)[:3])
        if str(df[col].dtype)[:3] == 'flo':

            for index in range(df.shape[0]):
                cl = ws.cell(row=index + 2, column=itr)
                cl.number_format = '0.00%'
                cl.value = float(df.iloc[index, itr - 1])

        if str(df[col].dtype)[:3] == 'int':

            for index in range(df.shape[0]):
                ws.cell(row=index + 2, column=itr).value = int(df.iloc[index, itr - 1])

        else:
            for index in range(df.shape[0]):
                cl = ws.cell(row=index + 2, column=itr)
                cl.value = (df.iloc[index, itr - 1])
                if itr == df.shape[1]:
                    if cl.value == 'К':
                        cl.fill = PatternFill(start_color=ck, end_color=ck, fill_type="solid")
                    elif cl.value == 'З':
                        cl.fill = PatternFill(start_color=cz, end_color=cz, fill_type="solid")
                    elif cl.value == 'Ж':
                        cl.fill = PatternFill(start_color=cs, end_color=cs, fill_type="solid")

                        #    добавляем границы
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            cl = ws.cell(row=i + 2, column=j + 1)
            cl.border = Border(top=Side(border_style='thin', color='ff000000'),
                               right=Side(border_style='thin', color='ff000000'),
                               bottom=Side(border_style='thin', color='ff000000'),
                               left=Side(border_style='thin', color='ff000000'))
            cl.alignment = Alignment(horizontal='center')

    # подравниваем по ширине значений

    coll = 1
    while coll < df.shape[1]+1:
        wdth = 5
        for i in range(2, df.shape[0] + 2):
            vl = ws.cell(row=i, column=coll).value
            if 'float' in str(type(vl)):
                vl = round(vl, 2)
            ms = len(str(vl))
            if ms > wdth:
                wdth = ms
        ws.column_dimensions[gkl(coll)].width = wdth + 5
        coll += 1

    # раздвигаем заголовки

    for i in range(1,df.shape[1]+1):
        cl = ws.cell(row=1, column=i)
        cl.alignment = Alignment(vertical='top', wrap_text=True)

#вставляет обычнй датафрейм на лист без форматирования
def tbl_to_sheet(df, ws, anch=None):

    if anch == None:
        anch = [ws.max_row + 4, 0]

    itr = 1
    for col in df.columns:
        itr += 1
        hd = ws.cell(row=anch[0] + 1, column=anch[1] + itr)
        hd.value = (col)
        hd.font = Font(bold=True)  # , color='ffffff')
        #   форматируем численные значения
        if str(df[col].dtype)[:3] == 'flo':

            for indx in range(df.shape[0]):
                cl = ws.cell(row=anch[0] + indx + 2, column=anch[1] + itr)
                cl.number_format = '0.00%'
                cl.value = float(df.iloc[indx, itr - 2])



        else:

            for indx2 in range(df.shape[0]):
                ws.cell(row=anch[0] + indx2 + 2, column=anch[1] + itr).value = df.iloc[indx2, itr - 2]
    k = 0
    for ind_v in df.index:
        ws.cell(row=anch[0] + k + 2, column=anch[1] + 1).value = str(ind_v)
        k += 1

    # добавляем границы
    for i in range(df.shape[0] + 1):
        for j in range(df.shape[1] + 1):
            cl = ws.cell(row=anch[0] + i + 1, column=anch[1] + j + 1)
            cl.border = Border(top=Side(border_style='thin', color='ff000000'),
                               right=Side(border_style='thin', color='ff000000'),
                               bottom=Side(border_style='thin', color='ff000000'),
                               left=Side(border_style='thin', color='ff000000'))
            cl.alignment = Alignment(horizontal='center')

    # подравниваем по ширине значений

    coll = 1
    while coll < df.shape[1] + 2:
        wdth = 0
        for i in range(2, df.shape[0] + 1):
            vl = ws.cell(row=anch[0] + i, column=anch[1] + coll ).value
            if 'float' in str(type(vl)):
                vl = round(vl, 2)
            ms = len(str(vl))
            if ms > wdth:
                wdth = ms
        ws.column_dimensions[gkl(coll)].width = wdth + 5
        coll += 1

    # раздвигаем заголовки

    for i in range(df.shape[1] + 1):
        cl = ws.cell(row=anch[0] + 1, column=anch[1] + i + 2)
        cl.alignment = Alignment(vertical='top', wrap_text=True)


# вставляет объект fig в книгу на лист в позицию
def insert_img(figr, ws, anchor_t='P1'):
    imdata = BytesIO()
    figr.savefig(imdata, format='png',bbox_inches='tight')#, pad_inches=2.5 )
    img = openpyxl.drawing.image.Image(imdata)
    img.anchor = anchor_t
    ws.add_image(img)



# функция выводящая тест в книгу
# при вставке новых картинок, старые удаляются. поэтому вставлять нужно все
# сразу и потом сохранять рабочую книгу.
def test_to_excel(test,wb,test_name=''):
    if test_name == '':
        test_name = test.tname
    print(f'{test_name}...')
    try:
        ws = wb.get_sheet_by_name(test_name)
        ws.delete_rows(1,ws.max_row)
    except:
        ws = wb.create_sheet(test_name)

    tbl_res_out(test.tbls[0],ws)
    for tdf in test.tbls[1:]:
        tbl_to_sheet(tdf,ws,[ws.max_row+5,0])
    i=ws.max_row + 10
    if len(test.grafls)>0:
        insert_img(test.grafls[0], ws)
        for fg in test.grafls[1:]:
            insert_img(fg,ws,'A'+str(i))
            plt.close(fg)
            i+=5

